import os
import streamlit as st
import sqlite3
import pandas as pd
from datetime import datetime
from contextlib import contextmanager
from zoneinfo import ZoneInfo
from io import BytesIO
import zipfile
import base64
import hashlib
import secrets
import html
import streamlit.components.v1 as components

try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# ===================== الثوابت =====================
# مسار قاعدة البيانات: من متغير بيئة (للنشر) أو محلي افتراضياً
DB_NAME = os.environ.get("DB_PATH", "taqyeem_system.db")

# إنشاء مجلد القاعدة إن كان مساره دليلاً غير موجود
_db_dir = os.path.dirname(DB_NAME)
if _db_dir and not os.path.exists(_db_dir):
    os.makedirs(_db_dir, exist_ok=True)

MAX_IMG_SIZE_MB = 20  # الحد الأعلى للصورة المرفوعة قبل الضغط
MAX_LOGIN_ATTEMPTS = 5
MIN_PASSWORD_LEN = 6

# المناطق الزمنية الشائعة في الشرق الأوسط وشمال أفريقيا
TIMEZONES = {
    "Asia/Riyadh": "🇸🇦 الرياض / الكويت / البحرين / قطر (UTC+3)",
    "Asia/Dubai": "🇦🇪 دبي / أبوظبي / مسقط (UTC+4)",
    "Africa/Cairo": "🇪🇬 القاهرة (UTC+2)",
    "Asia/Baghdad": "🇮🇶 بغداد (UTC+3)",
    "Asia/Amman": "🇯🇴 عمّان (UTC+3)",
    "Asia/Beirut": "🇱🇧 بيروت (UTC+3)",
    "Asia/Damascus": "🇸🇾 دمشق (UTC+3)",
    "Asia/Jerusalem": "🇵🇸 القدس (UTC+3)",
    "Asia/Aden": "🇾🇪 صنعاء (UTC+3)",
    "Africa/Khartoum": "🇸🇩 الخرطوم (UTC+2)",
    "Africa/Tripoli": "🇱🇾 طرابلس (UTC+2)",
    "Africa/Tunis": "🇹🇳 تونس (UTC+1)",
    "Africa/Algiers": "🇩🇿 الجزائر (UTC+1)",
    "Africa/Casablanca": "🇲🇦 الدار البيضاء (UTC+1)",
    "Asia/Tehran": "🇮🇷 طهران (UTC+3:30)",
    "Asia/Istanbul": "🇹🇷 إسطنبول (UTC+3)",
    "UTC": "🌍 UTC (التوقيت العالمي)",
}

# ===================== نظام الترجمة / Translation System =====================
LANGS = {
    "ar": {
        # Login
        "login_title": "🔐 تسجيل دخول النظام",
        "username": "اسم المستخدم",
        "username_ph": "أدخل اسم المستخدم",
        "password": "كلمة المرور",
        "password_ph": "أدخل كلمة المرور",
        "login_btn": "دخول",
        "login_failed": "⚠️ بيانات الدخول غير صحيحة",
        "attempts_left": "🔁 المحاولات المتبقية: {}",
        "max_attempts": "⛔ تم تجاوز الحد الأقصى للمحاولات ({}). أعد تشغيل التطبيق.",
        # Sidebar
        "lang_label": "🌐 اللغة",
        "sb_title": "🖼️ إعدادات التقرير",
        "logo_right": "شعار اليمين",
        "logo_left": "شعار اليسار",
        "org_name_label": "اسم الجهة / المؤسسة",
        "footer_text_label": "نص التذييل",
        "logout": "🚪 تسجيل الخروج",
        "admin_only": "⛔ هذه الصفحة للمدير فقط.",
        "welcome": "مرحباً بك يا {} ({})",
        # Menu
        "go_to": "انتقل إلى:",
        "m_dashboard": "📊 لوحة المؤشرات",
        "m_maintenance": "🛠️ الصيانة",
        "m_cleaning": "🧹 النظافة",
        "m_daily": "✅ المهام اليومية",
        "m_report_maint": "📋 تقرير بلاغ فردي",
        "m_report_clean": "🧽 تقرير نظافة فردي",
        "m_report_daily": "🧾 تقرير الجولة اليومية",
        "m_report_monthly": "📅 التقرير الشهري",
        "m_users": "👥 إدارة المستخدمين",
        # Dashboard
        "dashboard_h": "📊 حالة العمل الحالية",
        "metric_total": "إجمالي البلاغات",
        "metric_pending": "بانتظار الإصلاح",
        "metric_done": "تم الإصلاح",
        "metric_cleaning": "مهام النظافة",
        "metric_inspections": "الجولات التفقدية",
        # Maintenance
        "maint_h": "🛠️ إدارة مهام الصيانة",
        "tab_new_report": "📝 فتح بلاغ جديد",
        "tab_close_report": "🔧 إغلاق بلاغ معلق",
        "dept": "القسم",
        "location_label": "الموقع (رقم المكتب / الدور)",
        "problem_desc": "وصف المشكلة",
        "photo_before_label": "صورة العطل (قبل)",
        "submit_report": "إرسال البلاغ",
        "report_sent": "✅ تم إرسال البلاغ بنجاح",
        "fill_required": "يرجى تعبئة الموقع والوصف",
        "select_report": "اختر البلاغ المراد إغلاقه",
        "action_taken": "الإجراء المتخذ",
        "photo_after_label": "صورة الإنجاز (بعد)",
        "close_report": "إغلاق البلاغ",
        "report_closed": "✅ تم إغلاق البلاغ بنجاح",
        "no_pending": "لا توجد بلاغات معلقة حالياً.",
        # Cleaning
        "clean_h": "🧹 سجل النظافة اليومي",
        "tab_add_record": "📝 إضافة سجل",
        "tab_view_records": "📂 عرض السجلات",
        "clean_area": "منطقة التنظيف",
        "clean_type": "نوع التنظيف",
        "before_clean": "قبل التنظيف",
        "after_clean": "بعد التنظيف",
        "save_record": "حفظ السجل",
        "saved": "✨ تم الحفظ بنجاح",
        "enter_area": "يرجى إدخال منطقة التنظيف",
        "no_records": "لا توجد سجلات نظافة حالياً.",
        "view_photos": "🖼️ عرض الصور",
        "select_record_view": "اختر السجل لعرض صوره",
        "no_photo": "لا توجد صورة",
        "before_lbl": "**⚠️ قبل التنظيف**",
        "after_lbl": "**✅ بعد التنظيف**",
        # Daily Inspections
        "daily_h": "✅ الجولات اليومية للتفقّد",
        "tab_new_round": "📋 جولة جديدة",
        "tab_daily_log": "📂 السجل اليومي",
        "inspection_info": "💡 التقط صورة لكل بند تشيّك عليه — هذا يؤكّد أنك نفّذت المهمة فعلياً.",
        "general_notes": "ملاحظات عامة (اختياري)",
        "status_label": "الحالة",
        "notes_label": "ملاحظات",
        "check_photo": "📷 صورة التشييك",
        "save_inspection": "🚀 حفظ الجولة التفقدية",
        "no_item_checked": "⚠️ لم يتم تشييك أي بند. اختر حالة لبند واحد على الأقل.",
        "photo_required": "⚠️ يجب التقاط صورة لكل بند تم تشييكه. الصور الناقصة: {}",
        "inspection_saved": "✅ تم حفظ الجولة التفقدية ({} بند). رقم الجولة: DC-{}",
        "no_inspections": "لا توجد جولات تفقدية سابقة.",
        "view_details": "🔍 عرض تفاصيل جولة",
        "select_round": "اختر الجولة",
        # User Management
        "users_h": "👥 إدارة المستخدمين",
        "current_users": "المستخدمون الحاليون",
        "add_user_h": "➕ إضافة مستخدم جديد",
        "add_btn": "إضافة",
        "fill_all": "يرجى تعبئة جميع الحقول",
        "pwd_too_short": "⚠️ يجب أن تكون كلمة المرور {} أحرف على الأقل",
        "user_added": "✅ تم إضافة المستخدم '{}' بنجاح",
        "user_exists": "⚠️ اسم المستخدم موجود مسبقاً",
        "delete_user_h": "🗑️ حذف مستخدم",
        "select_to_delete": "اختر المستخدم للحذف",
        "delete_btn": "حذف",
        "user_deleted": "✅ تم حذف المستخدم '{}'",
        "no_other_users": "لا يوجد مستخدمون آخرون للحذف.",
        "change_pwd_h": "🔑 تغيير كلمة مرور مستخدم",
        "select_user": "اختر المستخدم",
        "new_pwd": "كلمة المرور الجديدة",
        "update_btn": "تحديث",
        "enter_new_pwd": "أدخل كلمة المرور الجديدة",
        "pwd_updated": "✅ تم تحديث كلمة مرور '{}'",
        "role": "الصلاحية",
        # Reports
        "report_single_h": "📋 تقرير بلاغ فردي",
        "no_reports": "لا توجد بلاغات مسجلة.",
        "select_report_id": "اختر رقم البلاغ",
        "download_html": "📥 تحميل التقرير (HTML)",
        "report_clean_h": "🧽 تقرير نظافة فردي",
        "no_clean_records": "لا توجد سجلات نظافة مسجّلة.",
        "select_record": "اختر السجل",
        "report_daily_h": "🧾 تقرير الجولة اليومية",
        "no_inspections_2": "لا توجد جولات تفقدية مسجّلة.",
        "report_monthly_h": "📅 التقرير الشهري",
        "from_date": "من تاريخ",
        "to_date": "إلى تاريخ",
        "generate_report": "🔍 استخراج التقرير",
        "download_monthly": "📥 تحميل التقرير الشهري (HTML)",
        "download_cleaning": "📥 تحميل تقرير النظافة (HTML)",
        "download_excel_zip": "📤 تصدير Excel + الصور (ZIP)",
        "excel_zip_info": "ℹ️ ملف ZIP يحتوي على ملف Excel + الصور في مجلدات مرقمة برقم الطلب",
        "to_official_caption": "ℹ️ لإصدار تقرير رسمي مع التوقيع، انتقل إلى تقرير نظافة فردي.",
        # Defaults
        "default_org": "وزارة الشؤون البلدية",
        "default_footer": "سري - للاستخدام الرسمي فقط",
        # Status display (for translating stored Arabic values)
        "disp_pending": "قيد الانتظار",
        "disp_done": "تم الإصلاح",
        "disp_ok": "سليم",
        "disp_needs_fix": "يحتاج صيانة",
        "disp_admin": "مدير",
        "disp_tech": "فني",
        # Assignment & Priority
        "assigned_to": "الفني المسؤول",
        "assign_to_label": "إسناد البلاغ إلى:",
        "no_assignment": "بدون إسناد",
        "priority_label": "الأولوية",
        "priority_urgent": "🔴 عاجل",
        "priority_normal": "🟢 عادي",
        "priority_low": "⚪ منخفض",
        # Custom date (backdating)
        "use_custom_date": "📅 استخدام تاريخ مخصّص (سابق)",
        "custom_date_label": "التاريخ",
        "custom_time_label": "الوقت",
        "backdate_info": "ℹ️ متاح للمدير والمشرف فقط — حد أقصى 30 يوم سابق",
        # Timezone
        "tz_h": "🕐 إعدادات الوقت والمنطقة الزمنية",
        "tz_label": "المنطقة الزمنية",
        "tz_current": "الوقت الحالي حسب الإعداد:",
        "tz_updated": "✅ تم تحديث المنطقة الزمنية إلى: {}",
        # Edit Role
        "edit_role_h": "🛡️ تعديل دور (صلاحيات) المستخدم",
        "current_role": "الدور الحالي:",
        "new_role_label": "الدور الجديد:",
        "role_updated": "✅ تم تحديث دور '{}' إلى '{}'",
        "cant_change_self": "⚠️ لا يمكنك تغيير دورك بنفسك (تجنّب فقدان صلاحيات المدير)",
        # Backup
        "backup_h": "💾 النسخ الاحتياطي",
        "backup_info": "ℹ️ احفظ نسخة من قاعدة البيانات بشكل دوري (أسبوعياً على الأقل) واحتفظ بها في مكان آمن.",
        "backup_btn": "📥 تحميل نسخة احتياطية الآن",
        "backup_size": "📊 حجم قاعدة البيانات الحالية:",
        "backup_records": "📈 إجمالي السجلات:",
        "backup_last_modified": "🕐 آخر تعديل:",
    },
    "en": {
        "login_title": "🔐 System Login",
        "username": "Username",
        "username_ph": "Enter username",
        "password": "Password",
        "password_ph": "Enter password",
        "login_btn": "Login",
        "login_failed": "⚠️ Invalid credentials",
        "attempts_left": "🔁 Attempts remaining: {}",
        "max_attempts": "⛔ Max login attempts ({}) exceeded. Please restart the app.",
        "lang_label": "🌐 Language",
        "sb_title": "🖼️ Report Settings",
        "logo_right": "Right Logo",
        "logo_left": "Left Logo",
        "org_name_label": "Organization Name",
        "footer_text_label": "Footer Text",
        "logout": "🚪 Logout",
        "admin_only": "⛔ Admin only.",
        "welcome": "Welcome, {} ({})",
        "go_to": "Go to:",
        "m_dashboard": "📊 Dashboard",
        "m_maintenance": "🛠️ Maintenance",
        "m_cleaning": "🧹 Cleaning",
        "m_daily": "✅ Daily Tasks",
        "m_report_maint": "📋 Maintenance Report",
        "m_report_clean": "🧽 Cleaning Report",
        "m_report_daily": "🧾 Daily Inspection Report",
        "m_report_monthly": "📅 Monthly Report",
        "m_users": "👥 User Management",
        "dashboard_h": "📊 Current Status",
        "metric_total": "Total Reports",
        "metric_pending": "Pending",
        "metric_done": "Completed",
        "metric_cleaning": "Cleaning Tasks",
        "metric_inspections": "Inspections",
        "maint_h": "🛠️ Maintenance Management",
        "tab_new_report": "📝 New Report",
        "tab_close_report": "🔧 Close Pending",
        "dept": "Department",
        "location_label": "Location (Office No. / Floor)",
        "problem_desc": "Problem Description",
        "photo_before_label": "Issue Photo (Before)",
        "submit_report": "Submit Report",
        "report_sent": "✅ Report submitted successfully",
        "fill_required": "Please fill in location and description",
        "select_report": "Select report to close",
        "action_taken": "Action Taken",
        "photo_after_label": "Completion Photo (After)",
        "close_report": "Close Report",
        "report_closed": "✅ Report closed successfully",
        "no_pending": "No pending reports.",
        "clean_h": "🧹 Daily Cleaning Log",
        "tab_add_record": "📝 Add Record",
        "tab_view_records": "📂 View Records",
        "clean_area": "Cleaning Area",
        "clean_type": "Cleaning Type",
        "before_clean": "Before Cleaning",
        "after_clean": "After Cleaning",
        "save_record": "Save Record",
        "saved": "✨ Saved successfully",
        "enter_area": "Please enter the cleaning area",
        "no_records": "No cleaning records yet.",
        "view_photos": "🖼️ View Photos",
        "select_record_view": "Select record to view photos",
        "no_photo": "No photo",
        "before_lbl": "**⚠️ Before Cleaning**",
        "after_lbl": "**✅ After Cleaning**",
        "daily_h": "✅ Daily Inspection Rounds",
        "tab_new_round": "📋 New Round",
        "tab_daily_log": "📂 Daily Log",
        "inspection_info": "💡 Take a photo of each item you check — this proves you actually performed the task.",
        "general_notes": "General Notes (Optional)",
        "status_label": "Status",
        "notes_label": "Notes",
        "check_photo": "📷 Inspection Photo",
        "save_inspection": "🚀 Save Inspection Round",
        "no_item_checked": "⚠️ No item checked. Choose a status for at least one item.",
        "photo_required": "⚠️ A photo is required for each checked item. Missing photos: {}",
        "inspection_saved": "✅ Inspection saved ({} items). Round ID: DC-{}",
        "no_inspections": "No previous inspection rounds.",
        "view_details": "🔍 View Round Details",
        "select_round": "Select round",
        "users_h": "👥 User Management",
        "current_users": "Current Users",
        "add_user_h": "➕ Add New User",
        "add_btn": "Add",
        "fill_all": "Please fill all fields",
        "pwd_too_short": "⚠️ Password must be at least {} characters",
        "user_added": "✅ User '{}' added successfully",
        "user_exists": "⚠️ Username already exists",
        "delete_user_h": "🗑️ Delete User",
        "select_to_delete": "Select user to delete",
        "delete_btn": "Delete",
        "user_deleted": "✅ User '{}' deleted",
        "no_other_users": "No other users to delete.",
        "change_pwd_h": "🔑 Change User Password",
        "select_user": "Select user",
        "new_pwd": "New Password",
        "update_btn": "Update",
        "enter_new_pwd": "Enter new password",
        "pwd_updated": "✅ Password for '{}' updated",
        "role": "Role",
        "report_single_h": "📋 Single Report",
        "no_reports": "No reports recorded.",
        "select_report_id": "Select report ID",
        "download_html": "📥 Download Report (HTML)",
        "report_clean_h": "🧽 Single Cleaning Report",
        "no_clean_records": "No cleaning records.",
        "select_record": "Select record",
        "report_daily_h": "🧾 Daily Inspection Report",
        "no_inspections_2": "No inspection rounds recorded.",
        "report_monthly_h": "📅 Monthly Report",
        "from_date": "From",
        "to_date": "To",
        "generate_report": "🔍 Generate Report",
        "download_monthly": "📥 Download Monthly Report (HTML)",
        "download_cleaning": "📥 Download Cleaning Report (HTML)",
        "download_excel_zip": "📤 Export Excel + Photos (ZIP)",
        "excel_zip_info": "ℹ️ ZIP file contains Excel + photos in folders named by report number",
        "to_official_caption": "ℹ️ For an official report with signatures, go to Single Cleaning Report.",
        "default_org": "Municipal Affairs Ministry",
        "default_footer": "Confidential - Official Use Only",
        "disp_pending": "Pending",
        "disp_done": "Completed",
        "disp_ok": "OK",
        "disp_needs_fix": "Needs Repair",
        "disp_admin": "Admin",
        "disp_tech": "Technician",
        "assigned_to": "Assigned To",
        "assign_to_label": "Assign report to:",
        "no_assignment": "Unassigned",
        "priority_label": "Priority",
        "priority_urgent": "🔴 Urgent",
        "priority_normal": "🟢 Normal",
        "priority_low": "⚪ Low",
        "use_custom_date": "📅 Use custom (past) date",
        "custom_date_label": "Date",
        "custom_time_label": "Time",
        "backdate_info": "ℹ️ Admin & Supervisor only — max 30 days back",
        "tz_h": "🕐 Time & Timezone Settings",
        "tz_label": "Timezone",
        "tz_current": "Current time per setting:",
        "tz_updated": "✅ Timezone updated to: {}",
        "edit_role_h": "🛡️ Edit User Role / Permissions",
        "current_role": "Current role:",
        "new_role_label": "New role:",
        "role_updated": "✅ Role of '{}' updated to '{}'",
        "cant_change_self": "⚠️ You cannot change your own role (to prevent losing admin access)",
        "backup_h": "💾 Backup",
        "backup_info": "ℹ️ Download a database backup periodically (at least weekly) and keep it in a safe location.",
        "backup_btn": "📥 Download Backup Now",
        "backup_size": "📊 Current database size:",
        "backup_records": "📈 Total records:",
        "backup_last_modified": "🕐 Last modified:",
    },
}


def t(key, *args):
    """Translation helper. Pass format args for placeholder strings."""
    if 'lang' not in st.session_state:
        st.session_state.lang = 'ar'
    lang = st.session_state.get('lang', 'ar')
    val = LANGS.get(lang, LANGS['ar']).get(key, key)
    if args:
        try:
            return val.format(*args)
        except (IndexError, KeyError):
            return val
    return val


def tr_status(value):
    """Translate stored Arabic status values to current language display."""
    mapping = {
        Status.PENDING: 'disp_pending',
        Status.DONE: 'disp_done',
        CheckStatus.OK: 'disp_ok',
        CheckStatus.NEEDS_FIX: 'disp_needs_fix',
    }
    return t(mapping.get(value, '')) if value in mapping else (value or '—')


# ===================== جدول ترجمة القيم المخزّنة =====================
DISPLAY_MAP_EN = {
    # بنود التفقّد / Inspection items
    "💡 الإنارة": "💡 Lighting",
    "❄️ المكيفات": "❄️ Air Conditioners",
    "🚻 دورات المياه": "🚻 Restrooms",
    "⚡ غرف الكهرباء": "⚡ Electrical Rooms",
    "🛗 المصاعد": "🛗 Elevators",
    "🚰 خزان الصرف الصحي": "🚰 Sewage Tank",
    "🌀 مراوح الشفط": "🌀 Exhaust Fans",
    # الحالات / Status values
    "قيد الانتظار": "Pending",
    "تم الإصلاح": "Completed",
    "سليم": "OK",
    "يحتاج صيانة": "Needs Repair",
    "—": "—",
    # الصلاحيات / Roles
    "مدير": "Admin",
    "مشرف": "Supervisor",
    "فني صيانة": "Maintenance Tech",
    "فني نظافة": "Cleaning Tech",
    "فني تفقّد": "Inspection Tech",
    "قارئ": "Reader",
    "فني": "Technician (Legacy)",
    # أقسام الصيانة / Maintenance departments
    "تكييف": "AC",
    "كهرباء": "Electrical",
    "سباكة": "Plumbing",
    "نجارة": "Carpentry",
    "أخرى": "Other",
    # أولويات الصيانة / Priorities
    "عاجل": "🔴 Urgent",
    "عادي": "🟢 Normal",
    "منخفض": "⚪ Low",
    # أنواع التنظيف / Cleaning types
    "يومي روتيني": "Daily Routine",
    "تنظيف عميق": "Deep Cleaning",
    "تلميع رخام": "Marble Polishing",
    "واجهات": "Facades",
}


def tr_display(value):
    """ترجمة القيم المخزّنة بالعربية لعرضها باللغة الحالية."""
    if value is None or value == "":
        return "—"
    if st.session_state.get('lang', 'ar') == 'en':
        return DISPLAY_MAP_EN.get(str(value), str(value))
    return str(value)


# بنود الجولة اليومية للتفقّد
CHECKLIST_ITEMS = [
    "💡 الإنارة",
    "❄️ المكيفات",
    "🚻 دورات المياه",
    "⚡ غرف الكهرباء",
    "🛗 المصاعد",
    "🚰 خزان الصرف الصحي",
    "🌀 مراوح الشفط",
]


class Status:
    PENDING = "قيد الانتظار"
    DONE = t("metric_done")


class Role:
    ADMIN = "مدير"
    SUPERVISOR = "مشرف"
    TECH_MAINT = "فني صيانة"
    TECH_CLEAN = "فني نظافة"
    TECH_INSPECT = "فني تفقّد"
    READER = "قارئ"
    TECH = "فني"  # قديم - للتوافق مع الحسابات الموجودة

# قائمة الأدوار القابلة للاختيار في النموذج الجديد (دون "فني" القديم)
SELECTABLE_ROLES = [
    "مدير",
    "مشرف",
    "فني صيانة",
    "فني نظافة",
    "قارئ",
]

# مصفوفة الصلاحيات لكل دور
# "ALL" تعني كل الصلاحيات
PERMISSIONS = {
    "مدير": "ALL",
    "مشرف": {
        "view_dashboard",
        "maint_open", "maint_close",
        "cleaning_add", "cleaning_view",
        "daily_new", "daily_view",
        "report_maint", "report_clean", "report_daily", "report_monthly",
    },
    "فني صيانة": {
        "maint_open", "maint_close",
        "daily_new", "daily_view",
    },
    "فني نظافة": {
        "cleaning_add", "cleaning_view",
    },
    # دور قديم - يحصل على نفس صلاحيات فني الصيانة للتوافق
    "فني تفقّد": {
        "maint_open", "maint_close",
        "daily_new", "daily_view",
    },
    "قارئ": {
        "view_dashboard",
        "cleaning_view",
        "daily_view",
        "report_maint", "report_clean", "report_daily", "report_monthly",
    },
    # توافق مع الحسابات القديمة بدور "فني"
    "فني": {
        "maint_open", "maint_close",
        "cleaning_add", "cleaning_view",
        "daily_new", "daily_view",
    },
}


class CheckStatus:
    OK = "سليم"
    NEEDS_FIX = "يحتاج صيانة"
    NOT_CHECKED = "—"


class Priority:
    URGENT = "عاجل"
    NORMAL = "عادي"
    LOW = "منخفض"

PRIORITIES = ["عاجل", "عادي", "منخفض"]


# ===================== إعدادات الصفحة =====================
st.set_page_config(
    page_title="نظام إدارة المرافق - تقييم",
    layout="centered",
    initial_sidebar_state="collapsed"
)


# ===================== دوال التشفير =====================

def hash_password(password: str) -> str:
    """تشفير كلمة المرور باستخدام PBKDF2-SHA256."""
    salt = secrets.token_hex(16)
    key = hashlib.pbkdf2_hmac(
        'sha256', password.encode('utf-8'), salt.encode('utf-8'), 200_000
    )
    return f"pbkdf2${salt}${key.hex()}"


def verify_password(password: str, stored: str) -> bool:
    """التحقق من كلمة المرور المشفرة."""
    if not stored:
        return False
    try:
        scheme, salt, key_hex = stored.split('$')
        if scheme != 'pbkdf2':
            return False
        key = hashlib.pbkdf2_hmac(
            'sha256', password.encode('utf-8'), salt.encode('utf-8'), 200_000
        )
        return secrets.compare_digest(key.hex(), key_hex)
    except (ValueError, AttributeError):
        return False


def is_hashed(password_field: str) -> bool:
    return isinstance(password_field, str) and password_field.startswith("pbkdf2$")


# ===================== الإعدادات والوقت =====================

def get_setting(key: str, default: str = "") -> str:
    """قراءة إعداد عام من قاعدة البيانات."""
    try:
        with get_db() as conn:
            row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
            return row["value"] if row else default
    except Exception:
        return default


def set_setting(key: str, value: str) -> None:
    """حفظ إعداد عام."""
    with get_db() as conn:
        conn.execute(
            "INSERT INTO settings (key, value) VALUES (?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value)
        )


def get_timezone() -> ZoneInfo:
    """الحصول على المنطقة الزمنية المضبوطة (افتراضياً الرياض)."""
    tz_name = get_setting("timezone", "Asia/Riyadh")
    try:
        return ZoneInfo(tz_name)
    except Exception:
        return ZoneInfo("Asia/Riyadh")


def now_local() -> datetime:
    """التاريخ والوقت الحالي بالمنطقة الزمنية المضبوطة."""
    return datetime.now(get_timezone())


# ===================== دوال مساعدة =====================

@contextmanager
def get_db():
    """مدير سياق آمن لاتصال قاعدة البيانات (commit/rollback تلقائي)."""
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def safe(text) -> str:
    """ينظّف النص من أي وسوم HTML قبل حقنه في القوالب لمنع XSS."""
    if text is None or text == "":
        return "—"
    try:
        if pd.isna(text):
            return "—"
    except (TypeError, ValueError):
        pass
    return html.escape(str(text))


def file_to_base64(uploaded_file, max_dim: int = 1280, quality: int = 80) -> str:
    """
    يحوّل الملف المرفوع إلى base64 مع ضغط ذكي.
    - يُصغّر أي صورة أكبر من max_dim × max_dim إلى الحدود (مع الحفاظ على النسب)
    - يحفظ كـ JPEG بجودة 80% (يوفر 70-90% من الحجم بدون فقدان دقة ملحوظ)
    """
    if uploaded_file is None:
        return ""
    data = uploaded_file.getvalue()

    # تحقق سريع من الحد الأقصى للأمان (20 ميجا)
    if len(data) > MAX_IMG_SIZE_MB * 1024 * 1024:
        st.warning(f"⚠️ حجم الصورة كبير جداً (أكبر من {MAX_IMG_SIZE_MB} ميجا). الرجاء استخدام صورة أصغر.")
        return ""

    # إذا Pillow غير متوفر، نخزن الصورة كما هي
    if not HAS_PIL:
        return base64.b64encode(data).decode()

    try:
        img = Image.open(BytesIO(data))

        # تصحيح اتجاه الصورة من EXIF (مهم لصور الجوّال)
        try:
            from PIL import ImageOps
            img = ImageOps.exif_transpose(img)
        except Exception:
            pass

        # تحويل لـ RGB (JPEG لا يدعم RGBA/P)
        if img.mode in ('RGBA', 'LA', 'P'):
            bg = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            bg.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = bg
        elif img.mode != 'RGB':
            img = img.convert('RGB')

        # تصغير الأبعاد إن كانت أكبر من max_dim
        if max(img.size) > max_dim:
            img.thumbnail((max_dim, max_dim), Image.LANCZOS)

        # حفظ كـ JPEG مضغوط
        buf = BytesIO()
        img.save(buf, format='JPEG', quality=quality, optimize=True, progressive=True)
        compressed = buf.getvalue()

        return base64.b64encode(compressed).decode()
    except Exception as e:
        st.warning(f"⚠️ خطأ في معالجة الصورة: {e}")
        # في حالة فشل المعالجة، نخزن الأصل لو حجمه معقول
        if len(data) <= 5 * 1024 * 1024:
            return base64.b64encode(data).decode()
        return ""


def render_img(b64: str, style: str, placeholder_html: str = None) -> str:
    """ينشئ وسم <img> من base64 أو عنصر بديل عند غياب الصورة."""
    if b64 and isinstance(b64, str) and b64.strip():
        return f'<img src="data:image/jpeg;base64,{b64}" style="{style}">'
    if placeholder_html is not None:
        return placeholder_html
    return '<div class="photo-placeholder">لا توجد صورة</div>'


def create_excel_zip(df_m, df_c, df_d_items, date_from, date_to) -> bytes:
    """ينشئ ملف ZIP يحتوي على Excel + الصور في مجلدات مرقمة برقم الطلب."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        st.error("⚠️ مكتبة openpyxl غير مثبتة")
        return b""

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        wb = Workbook()

        # ========== ورقة الصيانة ==========
        ws_m = wb.active
        ws_m.title = "الصيانة"
        try:
            ws_m.sheet_view.rightToLeft = True
        except Exception:
            pass
        ws_m.append([
            'الرقم', 'التاريخ', 'القسم', 'الموقع', 'الوصف', 'الحالة',
            'الأولوية', 'الفني المسؤول', 'الفني المنفّذ',
            'تاريخ الإغلاق', 'الإجراء المتخذ'
        ])
        for _, row in df_m.iterrows():
            rid = row['id']
            ws_m.append([
                f"TQ-{rid:04d}",
                row['date'] or '',
                row['dept'] or '',
                row['office_name'] or '',
                str(row['description'] or '')[:500],
                row['status'] or '',
                row['priority'] if 'priority' in row.index and row['priority'] else 'عادي',
                row['assigned_to'] if 'assigned_to' in row.index and row['assigned_to'] else '',
                row['tech_name'] or '',
                row['closed_date'] if 'closed_date' in row.index and row['closed_date'] else '',
                str(row['action_taken'] or '')[:500],
            ])
            # حفظ الصور
            if row['img_before']:
                try:
                    zf.writestr(f"Photos/Maintenance/TQ-{rid:04d}/before.jpg", base64.b64decode(row['img_before']))
                except Exception:
                    pass
            if row['img_after']:
                try:
                    zf.writestr(f"Photos/Maintenance/TQ-{rid:04d}/after.jpg", base64.b64decode(row['img_after']))
                except Exception:
                    pass

        # ========== ورقة النظافة ==========
        ws_c = wb.create_sheet("النظافة")
        try:
            ws_c.sheet_view.rightToLeft = True
        except Exception:
            pass
        ws_c.append(['الرقم', 'التاريخ', 'المنطقة', 'نوع التنظيف', 'المنفّذ'])
        for _, row in df_c.iterrows():
            rid = row['id']
            ws_c.append([
                f"CL-{rid:04d}",
                row['date'] or '',
                row['area'] or '',
                row['type'] or '',
                row['tech_name'] if 'tech_name' in row.index and row['tech_name'] else '',
            ])
            if row['img_before']:
                try:
                    zf.writestr(f"Photos/Cleaning/CL-{rid:04d}/before.jpg", base64.b64decode(row['img_before']))
                except Exception:
                    pass
            if row['img_after']:
                try:
                    zf.writestr(f"Photos/Cleaning/CL-{rid:04d}/after.jpg", base64.b64decode(row['img_after']))
                except Exception:
                    pass

        # ========== ورقة الجولات اليومية ==========
        ws_d = wb.create_sheet("الجولات_اليومية")
        try:
            ws_d.sheet_view.rightToLeft = True
        except Exception:
            pass
        ws_d.append(['رقم الجولة', 'التاريخ', 'المنفّذ', 'البند', 'الحالة', 'ملاحظات'])
        for _, row in df_d_items.iterrows():
            bid = row['batch_id']
            ws_d.append([
                f"DC-{bid}",
                row['date'] or '',
                row['tech_name'] or '',
                row['item'] or '',
                row['status'] or '',
                str(row['notes'] or '')[:500],
            ])
            if row['photo']:
                try:
                    # تنظيف اسم البند لاستخدامه في اسم الملف
                    item_clean = "".join(c for c in str(row['item'] or 'item') if c.isalnum() or c in (' ', '_', '-'))[:50] or "item"
                    item_clean = item_clean.strip().replace(' ', '_')
                    zf.writestr(
                        f"Photos/Daily_Checks/DC-{bid}/{item_clean}_{row['id']}.jpg",
                        base64.b64decode(row['photo'])
                    )
                except Exception:
                    pass

        # ========== تنسيق الأوراق ==========
        for ws in [ws_m, ws_c, ws_d]:
            # تنسيق رأس الجدول
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # تنسيق المحاذاة لكل البيانات
            for row_cells in ws.iter_rows(min_row=2):
                for cell in row_cells:
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            # ضبط عرض الأعمدة
            for col in ws.columns:
                max_length = 10
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_len = len(str(cell.value or ''))
                        if cell_len > max_length:
                            max_length = cell_len
                    except Exception:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 3, 60)

        # حفظ Excel داخل ZIP
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        zf.writestr(
            f"Monthly_Report_{date_from}_to_{date_to}.xlsx",
            excel_buffer.getvalue()
        )

        # إضافة ملف README بداخل الـ ZIP
        readme = f"""تقرير شهري
==================
الفترة: من {date_from} إلى {date_to}
تاريخ الإصدار: {now_local().strftime('%Y-%m-%d %H:%M')}

محتويات الملف:
- Monthly_Report_{date_from}_to_{date_to}.xlsx : ملف Excel فيه 3 أوراق (الصيانة + النظافة + الجولات اليومية)
- Photos/Maintenance/TQ-XXXX/ : صور كل بلاغ صيانة
- Photos/Cleaning/CL-XXXX/    : صور كل سجل نظافة
- Photos/Daily_Checks/DC-XX/  : صور كل جولة تفقد

نظام إدارة المرافق - تقييم
"""
        zf.writestr("README.txt", readme)

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def require_admin():
    """يوقف تنفيذ الصفحة إذا لم يكن المستخدم مديراً (تحقق خلفي حقيقي)."""
    if st.session_state.get('user_role') != Role.ADMIN:
        st.error(t("admin_only"))
        st.stop()


def has_permission(perm: str) -> bool:
    """تحقق هل المستخدم الحالي يملك صلاحية معينة."""
    role = st.session_state.get('user_role', '')
    perms = PERMISSIONS.get(role, set())
    if perms == "ALL":
        return True
    return perm in perms


def require_permission(perm: str):
    """يوقف الصفحة إذا لم يملك المستخدم الصلاحية."""
    if not has_permission(perm):
        st.error(t("admin_only"))
        st.stop()


# ===================== دالة التاريخ المخصص (للمدير والمشرف) =====================

def can_backdate() -> bool:
    """صلاحية تسجيل تاريخ سابق — المدير والمشرف فقط."""
    role = st.session_state.get('user_role', '')
    return role in (Role.ADMIN, Role.SUPERVISOR)


def custom_date_input(key_prefix: str, max_days_back: int = 30) -> str:
    """عرض حقل تاريخ مخصّص (للمدير والمشرف). يعيد سلسلة YYYY-MM-DD HH:MM."""
    if not can_backdate():
        return now_local().strftime("%Y-%m-%d %H:%M")

    use_custom = st.checkbox(t("use_custom_date"), key=f"{key_prefix}_toggle")
    if not use_custom:
        return now_local().strftime("%Y-%m-%d %H:%M")

    from datetime import timedelta as _td
    _now = now_local()
    min_date = (_now - _td(days=max_days_back)).date()
    max_date = _now.date()

    st.caption(t("backdate_info"))
    cols = st.columns(2)
    with cols[0]:
        d = st.date_input(
            t("custom_date_label"),
            value=max_date,
            min_value=min_date,
            max_value=max_date,
            key=f"{key_prefix}_date"
        )
    with cols[1]:
        tm = st.time_input(
            t("custom_time_label"),
            value=_now.time(),
            key=f"{key_prefix}_time"
        )
    # دمج التاريخ والوقت + إرفاق المنطقة الزمنية
    combined = datetime.combine(d, tm).replace(tzinfo=get_timezone())
    return combined.strftime("%Y-%m-%d %H:%M")


# ===================== تهيئة قاعدة البيانات =====================

def init_db():
    with get_db() as conn:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS users (
                       username TEXT PRIMARY KEY,
                       password TEXT,
                       role TEXT
                     )''')
        c.execute('''CREATE TABLE IF NOT EXISTS maintenance (
                       id INTEGER PRIMARY KEY AUTOINCREMENT,
                       date TEXT, dept TEXT, office_name TEXT, description TEXT,
                       status TEXT, img_before TEXT, img_after TEXT,
                       action_taken TEXT, tech_name TEXT
                     )''')
        c.execute('''CREATE TABLE IF NOT EXISTS cleaning (
                       id INTEGER PRIMARY KEY AUTOINCREMENT,
                       date TEXT, area TEXT, type TEXT,
                       img_before TEXT, img_after TEXT
                     )''')
        c.execute('''CREATE TABLE IF NOT EXISTS daily_checks (
                       id INTEGER PRIMARY KEY AUTOINCREMENT,
                       date TEXT,
                       batch_id TEXT,
                       tech_name TEXT,
                       item TEXT,
                       status TEXT,
                       photo TEXT,
                       notes TEXT
                     )''')
        # جدول الإعدادات العامة (المنطقة الزمنية وغيرها)
        c.execute('''CREATE TABLE IF NOT EXISTS settings (
                       key TEXT PRIMARY KEY,
                       value TEXT
                     )''')

        # ترقية: إضافة عمود tech_name لجدول النظافة إن لم يكن موجوداً
        try:
            c.execute("ALTER TABLE cleaning ADD COLUMN tech_name TEXT")
        except sqlite3.OperationalError:
            pass
        # ترقية: إضافة عمود closed_date لجدول الصيانة إن لم يكن موجوداً
        try:
            c.execute("ALTER TABLE maintenance ADD COLUMN closed_date TEXT")
        except sqlite3.OperationalError:
            pass
        # ترقية: إضافة عمود assigned_to لإسناد البلاغ لفني محدد
        try:
            c.execute("ALTER TABLE maintenance ADD COLUMN assigned_to TEXT")
        except sqlite3.OperationalError:
            pass
        # ترقية: إضافة عمود priority لتصنيف الأولوية
        try:
            c.execute("ALTER TABLE maintenance ADD COLUMN priority TEXT DEFAULT 'عادي'")
        except sqlite3.OperationalError:
            pass

        # إنشاء المستخدمين الافتراضيين مع كلمات مرور مشفّرة (مرة واحدة فقط)
        c.execute("SELECT username FROM users WHERE username='admin'")
        if not c.fetchone():
            c.execute(
                "INSERT INTO users VALUES (?,?,?)",
                ('admin', hash_password('admin123'), Role.ADMIN)
            )
        c.execute("SELECT username FROM users WHERE username='tech'")
        if not c.fetchone():
            c.execute(
                "INSERT INTO users VALUES (?,?,?)",
                ('tech', hash_password('tech123'), Role.TECH)
            )


init_db()

# ===================== نظام تسجيل الدخول =====================

# ===================== مفتاح تبديل اللغة / Language Toggle =====================
if 'lang' not in st.session_state:
    st.session_state.lang = 'ar'

_lang_options = {"🇸🇦 العربية": "ar", "🇬🇧 English": "en"}
_current_display = "🇸🇦 العربية" if st.session_state.lang == 'ar' else "🇬🇧 English"
_new_display = st.sidebar.selectbox(
    "🌐 اللغة / Language",
    options=list(_lang_options.keys()),
    index=list(_lang_options.keys()).index(_current_display),
    key="lang_selector"
)
_new_lang = _lang_options[_new_display]
if _new_lang != st.session_state.lang:
    st.session_state.lang = _new_lang
    st.rerun()

if 'login_attempts' not in st.session_state:
    st.session_state.login_attempts = 0

if 'logged_in' not in st.session_state:
    st.markdown(
        f"<h2 style='text-align:center;'>{t('login_title')}</h2>",
        unsafe_allow_html=True
    )

    if st.session_state.login_attempts >= MAX_LOGIN_ATTEMPTS:
        st.error(t("max_attempts", MAX_LOGIN_ATTEMPTS))
        st.stop()

    u = st.text_input(t("username"), placeholder=t("username_ph"))
    p = st.text_input(t("password"), type="password", placeholder=t("password_ph"))

    if st.session_state.login_attempts > 0:
        remaining = MAX_LOGIN_ATTEMPTS - st.session_state.login_attempts
        st.caption(t("attempts_left", remaining))

    if st.button(t("login_btn"), use_container_width=True):
        ok = False
        role = None
        with get_db() as conn:
            row = conn.execute(
                "SELECT username, password, role FROM users WHERE username=?",
                (u,)
            ).fetchone()
            if row is not None:
                stored = row['password']
                if is_hashed(stored):
                    if verify_password(p, stored):
                        ok = True
                        role = row['role']
                else:
                    # توافق مع البيانات القديمة (نص صريح) — يقبل ثم يُرحّل إلى مشفّر
                    if p == stored:
                        ok = True
                        role = row['role']
                        conn.execute(
                            "UPDATE users SET password=? WHERE username=?",
                            (hash_password(p), u)
                        )

        if ok:
            st.session_state.update({
                'logged_in': True,
                'user_role': role,
                'username': u,
                'login_attempts': 0,
            })
            st.rerun()
        else:
            st.session_state.login_attempts += 1
            st.error(t("login_failed"))
    st.stop()

# ===================== الشريط الجانبي =====================

st.sidebar.title(t("sb_title"))

# الشعارات: تُحفظ في الجلسة حتى لا تضيع عند كل rerun
l_logo_file = st.sidebar.file_uploader(t("logo_right"), type=['png', 'jpg', 'jpeg'], key='l_logo')
r_logo_file = st.sidebar.file_uploader(t("logo_left"), type=['png', 'jpg', 'jpeg'], key='r_logo')

if l_logo_file is not None:
    st.session_state.l_logo_b64 = base64.b64encode(l_logo_file.getvalue()).decode()
if r_logo_file is not None:
    st.session_state.r_logo_b64 = base64.b64encode(r_logo_file.getvalue()).decode()

l_logo_b64 = st.session_state.get('l_logo_b64', '')
r_logo_b64 = st.session_state.get('r_logo_b64', '')

org_name      = st.sidebar.text_input("اسم الجهة / المؤسسة", value=t("default_org"))
report_footer = st.sidebar.text_input(t("footer_text_label"), value=t("default_footer"))

is_admin = st.session_state.user_role == Role.ADMIN

# القائمة الديناميكية حسب الصلاحيات الفعلية
menu = []
if has_permission("view_dashboard"):
    menu.append(t("m_dashboard"))
if has_permission("maint_open") or has_permission("maint_close"):
    menu.append(t("m_maintenance"))
if has_permission("cleaning_add") or has_permission("cleaning_view"):
    menu.append(t("m_cleaning"))
if has_permission("daily_new") or has_permission("daily_view"):
    menu.append(t("m_daily"))
if has_permission("report_maint"):
    menu.append(t("m_report_maint"))
if has_permission("report_clean"):
    menu.append(t("m_report_clean"))
if has_permission("report_daily"):
    menu.append(t("m_report_daily"))
if has_permission("report_monthly"):
    menu.append(t("m_report_monthly"))
if is_admin:
    menu.append(t("m_users"))

# في حالة المستخدم بلا صلاحيات
if not menu:
    menu = ["(لا توجد صلاحيات)"]

choice = st.selectbox(t("go_to"), menu)

# ===================== لوحة المؤشرات (مدير فقط) =====================

if choice == t("m_dashboard"):
    require_permission("view_dashboard")
    st.header(t("dashboard_h"))
    with get_db() as conn:
        m_count = conn.execute("SELECT COUNT(*) FROM maintenance").fetchone()[0]
        c_count = conn.execute("SELECT COUNT(*) FROM cleaning").fetchone()[0]
        pending = conn.execute(
            "SELECT COUNT(*) FROM maintenance WHERE status=?", (Status.PENDING,)
        ).fetchone()[0]
        done = conn.execute(
            "SELECT COUNT(*) FROM maintenance WHERE status=?", (Status.DONE,)
        ).fetchone()[0]
        dc_batches = conn.execute(
            "SELECT COUNT(DISTINCT batch_id) FROM daily_checks"
        ).fetchone()[0]
        # البلاغات المتأخرة (3+ أيام بدون إغلاق)
        from datetime import timedelta as _td
        _cutoff = (now_local() - _td(days=3)).strftime("%Y-%m-%d %H:%M")
        late_reports = pd.read_sql_query(
            "SELECT id, date, office_name, dept, priority, assigned_to FROM maintenance "
            "WHERE status=? AND date < ? ORDER BY date ASC",
            conn, params=(Status.PENDING, _cutoff)
        )

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric(t("metric_total"), m_count)
    col2.metric(t("metric_pending"), pending, delta=f"-{len(late_reports)} متأخر" if len(late_reports) > 0 else None, delta_color="inverse")
    col3.metric(t("metric_done"), done)
    col4.metric(t("metric_cleaning"), c_count)
    col5.metric(t("metric_inspections"), dc_batches)

    # قسم تنبيه البلاغات المتأخرة
    if not late_reports.empty:
        st.divider()
        st.error(f"⚠️ يوجد **{len(late_reports)}** بلاغ متأخر (أكثر من 3 أيام بدون إغلاق)")
        with st.expander("📋 عرض البلاغات المتأخرة", expanded=True):
            for _, _r in late_reports.iterrows():
                _pri = _r['priority'] or 'عادي'
                _pri_emoji = '🔴' if _pri == 'عاجل' else ('⚪' if _pri == 'منخفض' else '🟢')
                _ass = _r['assigned_to'] or 'بدون إسناد'
                # حساب عدد الأيام
                try:
                    _date_obj = datetime.strptime(_r['date'], "%Y-%m-%d %H:%M")
                    _days = (now_local().replace(tzinfo=None) - _date_obj).days
                except Exception:
                    _days = '?'
                st.markdown(
                    f"- {_pri_emoji} **TQ-{_r['id']:04d}** | {tr_display(_r['dept'])} | "
                    f"{_r['office_name']} | منذ **{_days}** يوم | 👤 {_ass}"
                )

    st.divider()
    st.info(t("welcome", st.session_state.username, st.session_state.user_role))

# ===================== قسم الصيانة =====================

elif choice == t("m_maintenance"):
    st.header(t("maint_h"))
    t1, t2 = st.tabs([t("tab_new_report"), t("tab_close_report")])

    with t1:
        if not has_permission("maint_open"):
            st.info(t("admin_only"))
        else:
            # حقل التاريخ المخصص (خارج النموذج لأن checkbox يحتاج rerun)
            _maint_date = custom_date_input("maint_open")

            # جلب قائمة الفنيين المتاحين للإسناد (المدير + المشرف فقط يقدرون يسندون)
            _can_assign = st.session_state.get('user_role') in (Role.ADMIN, Role.SUPERVISOR)
            _technicians = []
            if _can_assign:
                with get_db() as conn:
                    _techs_df = pd.read_sql_query(
                        "SELECT username FROM users WHERE role IN (?, ?, ?, ?) ORDER BY username",
                        conn, params=(Role.TECH_MAINT, Role.TECH, "فني صيانة", "فني")
                    )
                _technicians = _techs_df['username'].tolist()

            with st.form("add_maintenance"):
                dept  = st.selectbox(t("dept"), ["تكييف", "كهرباء", "سباكة", "نجارة", "أخرى"], format_func=tr_display)
                loc   = st.text_input(t("location_label"))
                desc  = st.text_area(t("problem_desc"))
                # تصنيف الأولوية
                priority = st.selectbox(
                    t("priority_label"),
                    PRIORITIES,
                    index=1,  # افتراضياً: عادي
                    format_func=tr_display
                )
                # إسناد الفني (المدير/المشرف فقط)
                assigned = None
                if _can_assign and _technicians:
                    assigned = st.selectbox(
                        t("assign_to_label"),
                        [t("no_assignment")] + _technicians
                    )
                    if assigned == t("no_assignment"):
                        assigned = None
                img_b = st.file_uploader(t("photo_before_label"), type=['jpg', 'png', 'jpeg'])
                if st.form_submit_button(t("submit_report"), use_container_width=True):
                    if loc and desc:
                        with get_db() as conn:
                            conn.execute(
                                "INSERT INTO maintenance (date,dept,office_name,description,status,img_before,priority,assigned_to) "
                                "VALUES (?,?,?,?,?,?,?,?)",
                                (
                                    _maint_date,
                                    dept, loc, desc, Status.PENDING, file_to_base64(img_b),
                                    priority, assigned
                                )
                            )
                        st.success(t("report_sent"))
                    else:
                        st.warning(t("fill_required"))

    with t2:
        if not has_permission("maint_close"):
            st.info(t("admin_only"))
        else:
            with get_db() as conn:
                pending_tasks = pd.read_sql_query(
                    "SELECT id, office_name, description, priority, assigned_to FROM maintenance WHERE status=?",
                    conn, params=(Status.PENDING,)
                )
            # فلتر للفنيين: يشوفون فقط البلاغات المسندة لهم + غير المسندة
            _current_user = st.session_state.get('username', '')
            _user_role = st.session_state.get('user_role', '')
            if _user_role in (Role.TECH_MAINT, Role.TECH):
                pending_tasks = pending_tasks[
                    (pending_tasks['assigned_to'].isna()) |
                    (pending_tasks['assigned_to'] == _current_user)
                ]
            if not pending_tasks.empty:
                def _format_option(row):
                    pri = row['priority'] or 'عادي'
                    pri_emoji = '🔴' if pri == 'عاجل' else ('⚪' if pri == 'منخفض' else '🟢')
                    assigned = f" [👤 {row['assigned_to']}]" if row['assigned_to'] else " [بدون إسناد]"
                    return f"{pri_emoji} #{row['id']} - {row['office_name']}{assigned}"
                options = {
                    _format_option(row): row['id']
                    for _, row in pending_tasks.iterrows()
                }
                selected = st.selectbox(t("select_report"), list(options.keys()))
                task_id  = options[selected]
                # حقل تاريخ مخصّص للإغلاق
                _close_date = custom_date_input("maint_close")
                with st.form("close_task"):
                    action = st.text_area(t("action_taken"))
                    img_a  = st.file_uploader(t("photo_after_label"), type=['jpg', 'png', 'jpeg'])
                    if st.form_submit_button(t("close_report"), use_container_width=True):
                        with get_db() as conn:
                            conn.execute(
                                "UPDATE maintenance "
                                "SET status=?, action_taken=?, img_after=?, tech_name=?, closed_date=? "
                                "WHERE id=?",
                                (
                                    Status.DONE, action, file_to_base64(img_a),
                                    st.session_state.username, _close_date, task_id
                                )
                            )
                        st.success(t("report_closed"))
                        st.rerun()
            else:
                st.info(t("no_pending"))

# ===================== قسم النظافة =====================

elif choice == t("m_cleaning"):
    st.header(t("clean_h"))
    t1, t2 = st.tabs([t("tab_add_record"), t("tab_view_records")])

    with t1:
        if not has_permission("cleaning_add"):
            st.info(t("admin_only"))
        else:
            _clean_date = custom_date_input("clean_add")
            with st.form("cleaning_form"):
                area    = st.text_input(t("clean_area"))
                c_type  = st.selectbox(
                    t("clean_type"),
                    ["يومي روتيني", "تنظيف عميق", "تلميع رخام", "واجهات"],
                    format_func=tr_display,
                )
                c_img_b = st.file_uploader(t("before_clean"), type=['jpg', 'png', 'jpeg'])
                c_img_a = st.file_uploader(t("after_clean"),  type=['jpg', 'png', 'jpeg'])
                if st.form_submit_button(t("save_record"), use_container_width=True):
                    if area:
                        with get_db() as conn:
                            conn.execute(
                                "INSERT INTO cleaning (date,area,type,img_before,img_after,tech_name) "
                                "VALUES (?,?,?,?,?,?)",
                                (
                                    _clean_date,
                                    area, c_type,
                                    file_to_base64(c_img_b), file_to_base64(c_img_a),
                                    st.session_state.username
                                )
                            )
                        st.success(t("saved"))
                    else:
                        st.warning(t("enter_area"))

    with t2:
        if not has_permission("cleaning_view"):
            st.info(t("admin_only"))
        else:
            with get_db() as conn:
                df_cl = pd.read_sql_query(
                    "SELECT id, date, area, type, tech_name, img_before, img_after "
                    "FROM cleaning ORDER BY id DESC", conn
                )
            if df_cl.empty:
                st.info(t("no_records"))
            else:
                display_df = df_cl[['id', 'date', 'area', 'type', 'tech_name']].copy()
                cols_ar = ['الرقم', 'التاريخ', 'المنطقة', 'النوع', 'المنفذ']
                cols_en = ['ID', 'Date', 'Area', 'Type', 'Executor']
                display_df.columns = cols_en if st.session_state.get('lang','ar')=='en' else cols_ar
                st.dataframe(display_df, use_container_width=True, hide_index=True)

                st.divider()
                st.subheader(t("view_photos"))
                options = {
                    f"#{row['id']} - {row['date']} - {row['area']}": row['id']
                    for _, row in df_cl.iterrows()
                }
                picked = st.selectbox(t("select_record_view"), list(options.keys()))
                rec_id = options[picked]
                r = df_cl[df_cl['id'] == rec_id].iloc[0]

                c1, c2 = st.columns(2)
                with c1:
                    st.markdown(t("before_lbl"))
                    if r['img_before']:
                        st.image(base64.b64decode(r['img_before']), use_container_width=True)
                    else:
                        st.info(t("no_photo"))
                with c2:
                    st.markdown(t("after_lbl"))
                    if r['img_after']:
                        st.image(base64.b64decode(r['img_after']), use_container_width=True)
                    else:
                        st.info(t("no_photo"))

                if is_admin:
                    st.caption("ℹ️ لإصدار تقرير رسمي مع التوقيع، انتقل إلى \"🧽 تقرير نظافة فردي\".")

# ===================== المهام اليومية (الجولات التفقدية) =====================

elif choice == t("m_daily"):
    st.header(t("daily_h"))
    t1, t2 = st.tabs([t("tab_new_round"), t("tab_daily_log")])

    with t1:
        if not has_permission("daily_new"):
            st.info(t("admin_only"))
        else:
            st.info(t("inspection_info"))
            # حقل التاريخ المخصّص للجولة (يُحفظ في session_state ليُستخدم داخل النموذج)
            _daily_custom = custom_date_input("daily_round")
            st.session_state['_daily_custom_date_str'] = _daily_custom

            with st.form("daily_inspection_form", clear_on_submit=False):
                general_notes = st.text_area(t("general_notes"), key="dc_general_notes")
                st.divider()

                inspection_data = {}
                for idx, item in enumerate(CHECKLIST_ITEMS):
                    st.markdown(f"#### {tr_display(item)}")
                    col_a, col_b = st.columns([1, 1])
                    with col_a:
                        status = st.selectbox(
                            t("status_label"),
                            [CheckStatus.NOT_CHECKED, CheckStatus.OK, CheckStatus.NEEDS_FIX],
                            format_func=tr_display,
                            key=f"dc_status_{idx}",
                        )
                        notes = st.text_input(t("notes_label"), key=f"dc_notes_{idx}")
                    with col_b:
                        photo = st.file_uploader(
                            t("check_photo"),
                            type=['jpg', 'png', 'jpeg'],
                            key=f"dc_photo_{idx}"
                        )
                    inspection_data[item] = {
                        'status': status,
                        'photo': photo,
                        'notes': notes
                    }
                    st.divider()

                submitted = st.form_submit_button(
                    t("save_inspection"),
                    use_container_width=True
                )

                if submitted:
                    checked = {k: v for k, v in inspection_data.items()
                               if v['status'] != CheckStatus.NOT_CHECKED}

                    if not checked:
                        st.warning(t("no_item_checked"))
                    else:
                        missing_photos = [k for k, v in checked.items() if v['photo'] is None]
                        if missing_photos:
                            st.error(
                                "⚠️ يجب التقاط صورة لكل بند تم تشييكه. "
                                f"الصور الناقصة: {'، '.join(missing_photos)}"
                            )
                        else:
                            batch_id = now_local().strftime("%Y%m%d_%H%M%S")
                            ts = now_local().strftime("%Y-%m-%d %H:%M")
                            with get_db() as conn:
                                for item, d in checked.items():
                                    conn.execute(
                                        "INSERT INTO daily_checks "
                                        "(date, batch_id, tech_name, item, status, photo, notes) "
                                        "VALUES (?,?,?,?,?,?,?)",
                                        (
                                            ts, batch_id,
                                            st.session_state.username,
                                            item, d['status'],
                                            file_to_base64(d['photo']),
                                            d['notes'] or general_notes
                                        )
                                    )
                            st.success(
                                f"✅ تم حفظ الجولة التفقدية ({len(checked)} بند). "
                                f"رقم الجولة: DC-{batch_id}"
                            )

    with t2:
        if not has_permission("daily_view"):
            st.info(t("admin_only"))
        else:
            with get_db() as conn:
                batches = pd.read_sql_query(
                    "SELECT batch_id, MIN(date) as date, tech_name, COUNT(*) as items_count "
                    "FROM daily_checks GROUP BY batch_id ORDER BY batch_id DESC",
                    conn
                )

            if batches.empty:
                st.info(t("no_inspections"))
            else:
                display_b = batches.copy()
                display_b['batch_id'] = "DC-" + display_b['batch_id'].astype(str)
                cols_ar = ['رقم الجولة', 'التاريخ', 'المنفذ', 'عدد البنود']
                cols_en = ['Round ID', 'Date', 'Executor', 'Items Count']
                display_b.columns = cols_en if st.session_state.get('lang','ar')=='en' else cols_ar
                st.dataframe(display_b, use_container_width=True, hide_index=True)

                st.divider()
                st.subheader(t("view_details"))
                batch_options = {
                    f"DC-{row['batch_id']} | {row['date']} | {row['tech_name']}": row['batch_id']
                    for _, row in batches.iterrows()
                }
                picked = st.selectbox(t("select_round"), list(batch_options.keys()))
                batch_id = batch_options[picked]

                with get_db() as conn:
                    items_df = pd.read_sql_query(
                        "SELECT * FROM daily_checks WHERE batch_id=? ORDER BY id",
                        conn, params=(batch_id,)
                    )

                for _, row in items_df.iterrows():
                    with st.container():
                        cols = st.columns([2, 3])
                        with cols[0]:
                            icon = "🟢" if row['status'] == CheckStatus.OK else "🟠"
                            st.markdown(f"### {tr_display(row['item'])}")
                            st.markdown(f"**" + t("status_label") + ":** {} {}".format(icon, tr_display(row['status'])))
                            if row['notes']:
                                st.markdown(f"**{t('notes_label')}:** {row['notes']}")
                        with cols[1]:
                            if row['photo']:
                                st.image(base64.b64decode(row['photo']), width=320)
                            else:
                                st.info(t("no_photo"))
                        st.divider()

# ===================== تقرير بلاغ فردي (مدير فقط) =====================

elif choice == t("m_report_maint"):
    require_permission("report_maint")
    st.header(t("m_report_maint"))
    with get_db() as conn:
        df = pd.read_sql_query("SELECT * FROM maintenance ORDER BY id DESC", conn)

    if df.empty:
        st.info(t("no_reports"))
    else:
        report_id = st.selectbox(t("select_report_id"), df['id'].tolist())
        r = df[df['id'] == report_id].iloc[0]

        img_style    = "width:100%;max-height:220px;object-fit:cover;border-radius:6px;border:1px solid #dde3ec;"
        placeholder  = '<div class="photo-placeholder">لا توجد صورة</div>'

        logo_left_html  = render_img(l_logo_b64, "height:75px;object-fit:contain;",
                                     '<div style="width:75px;"></div>')
        logo_right_html = render_img(r_logo_b64, "height:75px;object-fit:contain;",
                                     '<div style="width:75px;"></div>')
        status_color    = "#27ae60" if r['status'] == Status.DONE else "#e67e22"
        action_html     = safe(r['action_taken']) if r['action_taken'] else 'لم يتم تسجيل إجراء بعد'

        report_html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Amiri:wght@400;700&family=Cairo:wght@400;600;700&display=swap" rel="stylesheet">
<style>
  *{{margin:0;padding:0;box-sizing:border-box;}}
  body{{font-family:'Cairo',sans-serif;background:#ffffff;color:#1a1a2e;}}
  .page{{max-width:820px;margin:0 auto;background:#ffffff;border:1px solid #dde3ec;box-shadow:0 4px 24px rgba(0,0,0,0.10);}}
  .header{{background:#ffffff;padding:0;display:flex;flex-direction:column;border-bottom:1px solid #dde3ec;}}
  .header-top{{display:flex;align-items:center;justify-content:space-between;padding:20px 28px;gap:16px;}}
  .header-center{{text-align:center;flex:1;}}
  .header-center .org{{font-family:'Amiri',serif;font-size:17px;color:#c9aa5f;letter-spacing:1px;margin-bottom:6px;}}
  .header-center .title{{font-size:23px;font-weight:700;color:#1b3a6b;margin-bottom:6px;}}
  .header-center .ref{{font-size:12px;color:#5a6a82;background:#ffffff;display:inline-block;padding:3px 16px;border-radius:20px;border:1px solid #dde3ec;}}
  .header-stripe{{height:5px;background:linear-gradient(90deg,#1b3a6b,#c9aa5f,#e8c97a,#c9aa5f,#1b3a6b);}}
  .meta-bar{{background:#ffffff;border-bottom:1px solid #dde3ec;padding:10px 28px;display:flex;gap:28px;align-items:center;flex-wrap:wrap;}}
  .meta-item{{font-size:12px;color:#5a6a82;display:flex;align-items:center;gap:5px;}}
  .meta-item strong{{color:#1b3a6b;font-weight:600;}}
  .body{{padding:24px 28px;}}
  .info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:22px;}}
  .info-card{{background:#ffffff;border:1px solid #dde3ec;border-right:4px solid #c9aa5f;border-radius:6px;padding:11px 15px;}}
  .info-card .label{{font-size:10px;color:#8a96aa;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px;}}
  .info-card .value{{font-size:14px;font-weight:700;color:#1b3a6b;}}
  .section-title{{font-size:13px;font-weight:700;color:#1b3a6b;background:#ffffff;border-right:4px solid #c9aa5f;padding:8px 14px;border-radius:4px;margin:18px 0 10px;letter-spacing:.3px;}}
  .detail-box{{background:#ffffff;border:1px solid #dde3ec;border-radius:6px;padding:13px 16px;font-size:14px;line-height:1.9;color:#333d4d;margin-bottom:6px;min-height:48px;}}
  .photos-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:6px;}}
  .photo-box{{text-align:center;}}
  .photo-label{{font-size:11px;font-weight:600;color:#fff;margin-bottom:7px;display:inline-block;padding:3px 14px;border-radius:20px;}}
  .photo-label.before{{background:#e67e22;}}
  .photo-label.after{{background:#27ae60;}}
  .photo-placeholder{{width:100%;height:110px;background:#ffffff;display:flex;align-items:center;justify-content:center;color:#aab;border-radius:6px;border:2px dashed #cdd4e0;font-size:13px;}}
  .status-badge{{display:inline-block;padding:4px 16px;border-radius:20px;font-size:12px;font-weight:700;color:white;background:{status_color};letter-spacing:.3px;}}
  .divider{{border:none;border-top:1px solid #e8ecf2;margin:20px 0;}}
  .footer{{background:#ffffff;border-top:2px solid #dde3ec;padding:20px 28px;display:flex;justify-content:space-between;align-items:flex-end;}}
  .sig-box{{text-align:center;min-width:150px;}}
  .sig-label{{font-size:11px;color:#8a96aa;margin-bottom:22px;font-weight:600;}}
  .sig-line{{border-top:1.5px solid #1b3a6b;width:130px;margin:0 auto 5px;}}
  .sig-name{{font-size:12px;color:#1b3a6b;font-weight:700;}}
  .stamp{{border:2.5px solid #c9aa5f;border-radius:50%;width:88px;height:88px;display:flex;flex-direction:column;align-items:center;justify-content:center;color:#c9aa5f;font-size:10px;font-weight:700;text-align:center;transform:rotate(-10deg);line-height:1.5;background:#fff;box-shadow:0 2px 10px rgba(201,170,95,.15);}}
  .confidential{{text-align:center;font-size:10px;color:#aab;padding:8px;border-top:1px solid #e8ecf2;letter-spacing:2px;background:#ffffff;}}
</style>
</head>
<body>
<div class="page">
  <div class="header">
    <div class="header-top">
      {logo_right_html}
      <div class="header-center">
        <div class="org">{safe(org_name)}</div>
        <div class="title">تقرير صيانة فني</div>
        <div class="ref">الرقم المرجعي: TQ-{r['id']:04d}</div>
      </div>
      {logo_left_html}
    </div>
    <div class="header-stripe"></div>
  </div>
  <div class="meta-bar">
    <div class="meta-item">📅 <strong>تاريخ الإصدار:</strong> {now_local().strftime('%Y-%m-%d')}</div>
    <div class="meta-item">👤 <strong>أُعدَّ بواسطة:</strong> {safe(st.session_state.username)}</div>
    <div class="meta-item">📌 <strong>الحالة:</strong> <span class="status-badge">{safe(r['status'])}</span></div>
  </div>
  <div class="body">
    <div class="info-grid">
      <div class="info-card"><div class="label">تاريخ البلاغ</div><div class="value">{safe(r['date'])}</div></div>
      <div class="info-card"><div class="label">القسم الفني</div><div class="value">{safe(r['dept'])}</div></div>
      <div class="info-card"><div class="label">الموقع / رقم المكتب</div><div class="value">{safe(r['office_name'])}</div></div>
      <div class="info-card"><div class="label">الفني المنفذ</div><div class="value">{safe(r['tech_name'])}</div></div>
    </div>
    <hr class="divider">
    <div class="section-title">🔍 وصف المشكلة</div>
    <div class="detail-box">{safe(r['description'])}</div>
    <div class="section-title">✅ الإجراء المتخذ</div>
    <div class="detail-box">{action_html}</div>
    <hr class="divider">
    <div class="section-title">📷 التوثيق المصور</div>
    <div class="photos-grid">
      <div class="photo-box">
        <span class="photo-label before">⚠️ قبل الإصلاح</span><br>
        {render_img(r['img_before'], img_style, placeholder)}
      </div>
      <div class="photo-box">
        <span class="photo-label after">✅ بعد الإصلاح</span><br>
        {render_img(r['img_after'], img_style, placeholder)}
      </div>
    </div>
  </div>
  <div class="footer">
    <div class="sig-box">
      <div class="sig-label">الفني المنفذ</div>
      <div class="sig-line"></div>
      <div class="sig-name">{safe(r['tech_name']) if r['tech_name'] else '———'}</div>
    </div>
    <div class="stamp">تم<br>الاعتماد<br>✦</div>
    <div class="sig-box">
      <div class="sig-label">المشرف المسؤول</div>
      <div class="sig-line"></div>
      <div class="sig-name">———</div>
    </div>
  </div>
  <div class="confidential">{safe(report_footer)}</div>
</div>
</body>
</html>"""

        components.html(report_html, height=960, scrolling=True)
        st.write("")
        st.download_button(
            t("download_html"),
            report_html,
            file_name=f"Report_TQ-{report_id:04d}.html",
            mime="text/html",
            use_container_width=True
        )

# ===================== تقرير نظافة فردي (مدير فقط) =====================

elif choice == t("m_report_clean"):
    require_permission("report_clean")
    st.header(t("m_report_clean"))
    with get_db() as conn:
        df = pd.read_sql_query("SELECT * FROM cleaning ORDER BY id DESC", conn)

    if df.empty:
        st.info(t("no_clean_records"))
    else:
        options = {
            f"#{row['id']} - {row['date']} - {row['area']}": row['id']
            for _, row in df.iterrows()
        }
        selected = st.selectbox(t("select_record"), list(options.keys()))
        rec_id = options[selected]
        r = df[df['id'] == rec_id].iloc[0]

        img_style   = "width:100%;max-height:240px;object-fit:cover;border-radius:6px;border:1px solid #dde3ec;"
        placeholder = '<div class="photo-placeholder">لا توجد صورة</div>'

        logo_left_html  = render_img(l_logo_b64, "height:75px;object-fit:contain;",
                                     '<div style="width:75px;"></div>')
        logo_right_html = render_img(r_logo_b64, "height:75px;object-fit:contain;",
                                     '<div style="width:75px;"></div>')

        tech_value = r['tech_name'] if 'tech_name' in r.index else None

        clean_html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Amiri:wght@400;700&family=Cairo:wght@400;600;700&display=swap" rel="stylesheet">
<style>
  *{{margin:0;padding:0;box-sizing:border-box;}}
  body{{font-family:'Cairo',sans-serif;background:#ffffff;color:#1a1a2e;}}
  .page{{max-width:820px;margin:0 auto;background:#ffffff;border:1px solid #dde3ec;box-shadow:0 4px 24px rgba(0,0,0,0.10);}}
  .header{{background:#ffffff;display:flex;flex-direction:column;border-bottom:1px solid #dde3ec;}}
  .header-top{{display:flex;align-items:center;justify-content:space-between;padding:20px 28px;gap:16px;}}
  .header-center{{text-align:center;flex:1;}}
  .header-center .org{{font-family:'Amiri',serif;font-size:17px;color:#c9aa5f;letter-spacing:1px;margin-bottom:6px;}}
  .header-center .title{{font-size:23px;font-weight:700;color:#1b3a6b;margin-bottom:6px;}}
  .header-center .ref{{font-size:12px;color:#5a6a82;background:#ffffff;display:inline-block;padding:3px 16px;border-radius:20px;border:1px solid #dde3ec;}}
  .header-stripe{{height:5px;background:linear-gradient(90deg,#1b3a6b,#c9aa5f,#e8c97a,#c9aa5f,#1b3a6b);}}
  .meta-bar{{background:#ffffff;border-bottom:1px solid #dde3ec;padding:10px 28px;display:flex;gap:28px;align-items:center;flex-wrap:wrap;}}
  .meta-item{{font-size:12px;color:#5a6a82;display:flex;align-items:center;gap:5px;}}
  .meta-item strong{{color:#1b3a6b;font-weight:600;}}
  .body{{padding:24px 28px;}}
  .info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:22px;}}
  .info-card{{background:#ffffff;border:1px solid #dde3ec;border-right:4px solid #c9aa5f;border-radius:6px;padding:11px 15px;}}
  .info-card .label{{font-size:10px;color:#8a96aa;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px;}}
  .info-card .value{{font-size:14px;font-weight:700;color:#1b3a6b;}}
  .section-title{{font-size:13px;font-weight:700;color:#1b3a6b;background:#ffffff;border-right:4px solid #c9aa5f;padding:8px 14px;border-radius:4px;margin:18px 0 10px;}}
  .photos-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;}}
  .photo-box{{text-align:center;}}
  .photo-label{{font-size:11px;font-weight:600;color:#fff;margin-bottom:7px;display:inline-block;padding:3px 14px;border-radius:20px;}}
  .photo-label.before{{background:#e67e22;}}
  .photo-label.after{{background:#27ae60;}}
  .photo-placeholder{{width:100%;height:140px;background:#ffffff;display:flex;align-items:center;justify-content:center;color:#aab;border-radius:6px;border:2px dashed #cdd4e0;font-size:13px;}}
  .divider{{border:none;border-top:1px solid #e8ecf2;margin:20px 0;}}
  .footer{{background:#ffffff;border-top:2px solid #dde3ec;padding:20px 28px;display:flex;justify-content:space-between;align-items:flex-end;}}
  .sig-box{{text-align:center;min-width:150px;}}
  .sig-label{{font-size:11px;color:#8a96aa;margin-bottom:22px;font-weight:600;}}
  .sig-line{{border-top:1.5px solid #1b3a6b;width:130px;margin:0 auto 5px;}}
  .sig-name{{font-size:12px;color:#1b3a6b;font-weight:700;}}
  .stamp{{border:2.5px solid #c9aa5f;border-radius:50%;width:88px;height:88px;display:flex;flex-direction:column;align-items:center;justify-content:center;color:#c9aa5f;font-size:10px;font-weight:700;text-align:center;transform:rotate(-10deg);line-height:1.5;background:#fff;box-shadow:0 2px 10px rgba(201,170,95,.15);}}
  .confidential{{text-align:center;font-size:10px;color:#aab;padding:8px;border-top:1px solid #e8ecf2;letter-spacing:2px;background:#ffffff;}}
</style>
</head>
<body>
<div class="page">
  <div class="header">
    <div class="header-top">
      {logo_right_html}
      <div class="header-center">
        <div class="org">{safe(org_name)}</div>
        <div class="title">تقرير نظافة</div>
        <div class="ref">الرقم المرجعي: CL-{r['id']:04d}</div>
      </div>
      {logo_left_html}
    </div>
    <div class="header-stripe"></div>
  </div>
  <div class="meta-bar">
    <div class="meta-item">📅 <strong>تاريخ الإصدار:</strong> {now_local().strftime('%Y-%m-%d')}</div>
    <div class="meta-item">👤 <strong>أُعدَّ بواسطة:</strong> {safe(st.session_state.username)}</div>
  </div>
  <div class="body">
    <div class="info-grid">
      <div class="info-card"><div class="label">تاريخ التنظيف</div><div class="value">{safe(r['date'])}</div></div>
      <div class="info-card"><div class="label">المنطقة</div><div class="value">{safe(r['area'])}</div></div>
      <div class="info-card"><div class="label">نوع التنظيف</div><div class="value">{safe(r['type'])}</div></div>
      <div class="info-card"><div class="label">المنفذ</div><div class="value">{safe(tech_value)}</div></div>
    </div>
    <hr class="divider">
    <div class="section-title">📷 التوثيق المصور</div>
    <div class="photos-grid">
      <div class="photo-box">
        <span class="photo-label before">⚠️ قبل التنظيف</span><br>
        {render_img(r['img_before'], img_style, placeholder)}
      </div>
      <div class="photo-box">
        <span class="photo-label after">✅ بعد التنظيف</span><br>
        {render_img(r['img_after'], img_style, placeholder)}
      </div>
    </div>
  </div>
  <div class="footer">
    <div class="sig-box">
      <div class="sig-label">المنفذ</div>
      <div class="sig-line"></div>
      <div class="sig-name">{safe(tech_value) if tech_value else '———'}</div>
    </div>
    <div class="stamp">تم<br>الاعتماد<br>✦</div>
    <div class="sig-box">
      <div class="sig-label">المشرف المسؤول</div>
      <div class="sig-line"></div>
      <div class="sig-name">———</div>
    </div>
  </div>
  <div class="confidential">{safe(report_footer)}</div>
</div>
</body>
</html>"""

        components.html(clean_html, height=900, scrolling=True)
        st.write("")
        st.download_button(
            t("download_cleaning"),
            clean_html,
            file_name=f"Cleaning_CL-{rec_id:04d}.html",
            mime="text/html",
            use_container_width=True
        )

# ===================== تقرير الجولة اليومية (مدير فقط) =====================

elif choice == t("m_report_daily"):
    require_permission("report_daily")
    st.header(t("m_report_daily"))

    with get_db() as conn:
        batches = pd.read_sql_query(
            "SELECT batch_id, MIN(date) as date, tech_name, COUNT(*) as items_count "
            "FROM daily_checks GROUP BY batch_id ORDER BY batch_id DESC",
            conn
        )

    if batches.empty:
        st.info(t("no_inspections_2"))
    else:
        batch_options = {
            f"DC-{row['batch_id']} | {row['date']} | {row['tech_name']}": row['batch_id']
            for _, row in batches.iterrows()
        }
        selected = st.selectbox(t("select_round"), list(batch_options.keys()))
        batch_id = batch_options[selected]

        with get_db() as conn:
            items_df = pd.read_sql_query(
                "SELECT * FROM daily_checks WHERE batch_id=? ORDER BY id",
                conn, params=(batch_id,)
            )

        first_row = items_df.iloc[0]
        report_date = first_row['date']
        tech_name = first_row['tech_name']
        ok_count = len(items_df[items_df['status'] == CheckStatus.OK])
        fix_count = len(items_df[items_df['status'] == CheckStatus.NEEDS_FIX])

        items_html = ""
        for _, row in items_df.iterrows():
            sc = "#27ae60" if row['status'] == CheckStatus.OK else "#e67e22"
            photo_html = render_img(
                row['photo'],
                "width:100%;max-height:200px;object-fit:cover;border-radius:6px;border:1px solid #dde3ec;",
                '<div class="photo-placeholder">لا توجد صورة</div>'
            )
            items_html += f"""
            <div class="check-item">
              <div class="check-header">
                <div class="check-name">{safe(row['item'])}</div>
                <div class="check-status" style="background:{sc};">{safe(row['status'])}</div>
              </div>
              <div class="check-body">
                <div class="check-photo">{photo_html}</div>
                <div class="check-meta">
                  <div class="meta-line"><strong>الوقت:</strong> {safe(row['date'])}</div>
                  <div class="meta-line"><strong>ملاحظات:</strong> {safe(row['notes']) if row['notes'] else 'لا يوجد'}</div>
                </div>
              </div>
            </div>
            """

        logo_left_html  = render_img(l_logo_b64, "height:75px;object-fit:contain;",
                                     '<div style="width:75px;"></div>')
        logo_right_html = render_img(r_logo_b64, "height:75px;object-fit:contain;",
                                     '<div style="width:75px;"></div>')

        daily_html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Amiri:wght@400;700&family=Cairo:wght@400;600;700&display=swap" rel="stylesheet">
<style>
  *{{margin:0;padding:0;box-sizing:border-box;}}
  body{{font-family:'Cairo',sans-serif;background:#ffffff;color:#1a1a2e;}}
  .page{{max-width:820px;margin:0 auto;background:#ffffff;border:1px solid #dde3ec;box-shadow:0 4px 24px rgba(0,0,0,0.10);}}
  .header{{background:#ffffff;display:flex;flex-direction:column;border-bottom:1px solid #dde3ec;}}
  .header-top{{display:flex;align-items:center;justify-content:space-between;padding:20px 28px;gap:16px;}}
  .header-center{{text-align:center;flex:1;}}
  .header-center .org{{font-family:'Amiri',serif;font-size:17px;color:#c9aa5f;letter-spacing:1px;margin-bottom:6px;}}
  .header-center .title{{font-size:23px;font-weight:700;color:#1b3a6b;margin-bottom:6px;}}
  .header-center .ref{{font-size:12px;color:#5a6a82;background:#ffffff;display:inline-block;padding:3px 16px;border-radius:20px;border:1px solid #dde3ec;}}
  .header-stripe{{height:5px;background:linear-gradient(90deg,#1b3a6b,#c9aa5f,#e8c97a,#c9aa5f,#1b3a6b);}}
  .meta-bar{{background:#ffffff;border-bottom:1px solid #dde3ec;padding:10px 28px;display:flex;gap:28px;align-items:center;flex-wrap:wrap;}}
  .meta-item{{font-size:12px;color:#5a6a82;display:flex;align-items:center;gap:5px;}}
  .meta-item strong{{color:#1b3a6b;font-weight:600;}}
  .body{{padding:24px 28px;}}
  .stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:22px;}}
  .stat-card{{background:#ffffff;border:1px solid #dde3ec;border-top:4px solid #c9aa5f;border-radius:8px;padding:16px;text-align:center;}}
  .stat-card .num{{font-size:32px;font-weight:700;color:#1b3a6b;line-height:1;}}
  .stat-card .num.green{{color:#27ae60;}}
  .stat-card .num.orange{{color:#e67e22;}}
  .stat-card .lbl{{font-size:11px;color:#8a96aa;margin-top:6px;font-weight:600;}}
  .check-item{{border:1px solid #dde3ec;border-radius:8px;margin-bottom:12px;overflow:hidden;border-right:4px solid #c9aa5f;}}
  .check-header{{padding:12px 16px;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid #dde3ec;background:#ffffff;}}
  .check-name{{font-size:15px;font-weight:700;color:#1b3a6b;}}
  .check-status{{color:#fff;padding:4px 12px;border-radius:20px;font-size:11px;font-weight:700;}}
  .check-body{{padding:12px 16px;display:grid;grid-template-columns:1fr 1.5fr;gap:14px;align-items:center;}}
  .check-photo img{{max-width:100%;display:block;}}
  .photo-placeholder{{width:100%;height:120px;background:#ffffff;display:flex;align-items:center;justify-content:center;color:#aab;border-radius:6px;border:2px dashed #cdd4e0;font-size:13px;}}
  .check-meta{{font-size:13px;color:#5a6a82;line-height:1.9;}}
  .check-meta strong{{color:#1b3a6b;}}
  .meta-line{{margin-bottom:4px;}}
  .divider{{border:none;border-top:1px solid #e8ecf2;margin:18px 0;}}
  .footer{{background:#ffffff;border-top:2px solid #dde3ec;padding:20px 28px;display:flex;justify-content:space-between;align-items:flex-end;}}
  .sig-box{{text-align:center;min-width:150px;}}
  .sig-label{{font-size:11px;color:#8a96aa;margin-bottom:22px;font-weight:600;}}
  .sig-line{{border-top:1.5px solid #1b3a6b;width:130px;margin:0 auto 5px;}}
  .sig-name{{font-size:12px;color:#1b3a6b;font-weight:700;}}
  .stamp{{border:2.5px solid #c9aa5f;border-radius:50%;width:88px;height:88px;display:flex;flex-direction:column;align-items:center;justify-content:center;color:#c9aa5f;font-size:10px;font-weight:700;text-align:center;transform:rotate(-10deg);line-height:1.5;background:#fff;box-shadow:0 2px 10px rgba(201,170,95,.15);}}
  .confidential{{text-align:center;font-size:10px;color:#aab;padding:8px;border-top:1px solid #e8ecf2;letter-spacing:2px;background:#ffffff;}}
</style>
</head>
<body>
<div class="page">
  <div class="header">
    <div class="header-top">
      {logo_right_html}
      <div class="header-center">
        <div class="org">{safe(org_name)}</div>
        <div class="title">تقرير الجولة اليومية للتفقّد</div>
        <div class="ref">الرقم المرجعي: DC-{batch_id}</div>
      </div>
      {logo_left_html}
    </div>
    <div class="header-stripe"></div>
  </div>
  <div class="meta-bar">
    <div class="meta-item">📅 <strong>تاريخ الجولة:</strong> {safe(report_date)}</div>
    <div class="meta-item">👤 <strong>الفني المنفّذ:</strong> {safe(tech_name)}</div>
    <div class="meta-item">🕐 <strong>تاريخ الإصدار:</strong> {now_local().strftime('%Y-%m-%d %H:%M')}</div>
  </div>
  <div class="body">
    <div class="stats">
      <div class="stat-card"><div class="num">{len(items_df)}</div><div class="lbl">إجمالي البنود</div></div>
      <div class="stat-card"><div class="num green">{ok_count}</div><div class="lbl">سليم</div></div>
      <div class="stat-card"><div class="num orange">{fix_count}</div><div class="lbl">يحتاج صيانة</div></div>
    </div>
    <hr class="divider">
    {items_html}
  </div>
  <div class="footer">
    <div class="sig-box">
      <div class="sig-label">الفني المنفّذ</div>
      <div class="sig-line"></div>
      <div class="sig-name">{safe(tech_name)}</div>
    </div>
    <div class="stamp">تم<br>الاعتماد<br>✦</div>
    <div class="sig-box">
      <div class="sig-label">المشرف المسؤول</div>
      <div class="sig-line"></div>
      <div class="sig-name">———</div>
    </div>
  </div>
  <div class="confidential">{safe(report_footer)}</div>
</div>
</body>
</html>"""

        components.html(daily_html, height=1100, scrolling=True)
        st.write("")
        st.download_button(
            t("download_html"),
            daily_html,
            file_name=f"DailyCheck_DC-{batch_id}.html",
            mime="text/html",
            use_container_width=True
        )

# ===================== التقرير الشهري (مدير فقط) =====================

elif choice == t("m_report_monthly"):
    require_permission("report_monthly")
    st.header(t("m_report_monthly"))

    col1, col2 = st.columns(2)
    with col1:
        date_from = st.date_input(t("from_date"), value=datetime.now().replace(day=1))
    with col2:
        date_to = st.date_input(t("to_date"), value=datetime.now())

    if st.button(t("generate_report"), use_container_width=True):
        with get_db() as conn:
            df_m = pd.read_sql_query(
                "SELECT * FROM maintenance WHERE date >= ? AND date <= ? ORDER BY id ASC",
                conn,
                params=(date_from.strftime("%Y-%m-%d"), date_to.strftime("%Y-%m-%d 23:59"))
            )
            df_c = pd.read_sql_query(
                "SELECT * FROM cleaning WHERE date >= ? AND date <= ? ORDER BY id ASC",
                conn,
                params=(date_from.strftime("%Y-%m-%d"), date_to.strftime("%Y-%m-%d 23:59"))
            )
            df_d = pd.read_sql_query(
                "SELECT batch_id, MIN(date) as date, tech_name, COUNT(*) as items_count, "
                "SUM(CASE WHEN status=? THEN 1 ELSE 0 END) as ok_count, "
                "SUM(CASE WHEN status=? THEN 1 ELSE 0 END) as fix_count "
                "FROM daily_checks WHERE date >= ? AND date <= ? "
                "GROUP BY batch_id ORDER BY batch_id ASC",
                conn,
                params=(CheckStatus.OK, CheckStatus.NEEDS_FIX,
                        date_from.strftime("%Y-%m-%d"), date_to.strftime("%Y-%m-%d 23:59"))
            )
            # جلب البنود التفصيلية مع الصور لكل جولة
            df_d_items = pd.read_sql_query(
                "SELECT id, batch_id, date, tech_name, item, status, photo, notes "
                "FROM daily_checks WHERE date >= ? AND date <= ? ORDER BY batch_id, id",
                conn,
                params=(date_from.strftime("%Y-%m-%d"), date_to.strftime("%Y-%m-%d 23:59"))
            )

        total       = len(df_m)
        done_count  = len(df_m[df_m['status'] == Status.DONE])
        pend_count  = len(df_m[df_m['status'] == Status.PENDING])
        clean_count = len(df_c)
        daily_count = len(df_d)
        pct         = f"{(done_count/total*100):.0f}%" if total > 0 else "—"

        small_img_style = "width:65px;height:50px;object-fit:cover;border-radius:4px;"

        maint_rows = ""
        for _, row in df_m.iterrows():
            sc = "#27ae60" if row['status'] == Status.DONE else "#e67e22"
            ib = render_img(row['img_before'], small_img_style, "—")
            ia = render_img(row['img_after'], small_img_style, "—")
            desc_str = str(row['description']) if row['description'] else ''
            desc_short = desc_str[:45] + ('...' if len(desc_str) > 45 else '')
            maint_rows += f"<tr><td>TQ-{row['id']:04d}</td><td>{safe(row['date'])}</td><td>{safe(row['dept'])}</td><td>{safe(row['office_name'])}</td><td style='text-align:right;'>{safe(desc_short)}</td><td><span style='background:{sc};color:#fff;padding:2px 9px;border-radius:12px;font-size:11px;'>{safe(row['status'])}</span></td><td>{safe(row['tech_name'])}</td><td>{ib}</td><td>{ia}</td></tr>"

        clean_rows = ""
        for _, row in df_c.iterrows():
            cb = render_img(row['img_before'], small_img_style, "—")
            ca = render_img(row['img_after'], small_img_style, "—")
            tech = row['tech_name'] if 'tech_name' in row.index else None
            clean_rows += f"<tr><td>CL-{row['id']:04d}</td><td>{safe(row['date'])}</td><td>{safe(row['area'])}</td><td>{safe(row['type'])}</td><td>{safe(tech)}</td><td>{cb}</td><td>{ca}</td></tr>"

        # تجميع صور الجولات حسب batch_id
        photos_by_batch = {}
        for _, ph_row in df_d_items.iterrows():
            bid = ph_row['batch_id']
            if bid not in photos_by_batch:
                photos_by_batch[bid] = []
            if ph_row['photo']:
                photos_by_batch[bid].append((ph_row['item'], ph_row['status'], ph_row['photo']))

        daily_rows = ""
        thumb_style = "width:55px;height:42px;object-fit:cover;border-radius:4px;margin:1px;border:1px solid #dde3ec;"
        for _, row in df_d.iterrows():
            bid = row['batch_id']
            # بناء معرض الصور المصغّرة لهذه الجولة
            photos_html = ""
            for item_name, item_status, photo_b64 in photos_by_batch.get(bid, []):
                border_color = "#27ae60" if item_status == CheckStatus.OK else "#e67e22"
                photos_html += f'<img src="data:image/jpeg;base64,{photo_b64}" style="{thumb_style}border-color:{border_color};" title="{safe(item_name)} - {safe(item_status)}">'
            if not photos_html:
                photos_html = "—"
            daily_rows += f"<tr><td>DC-{safe(row['batch_id'])}</td><td>{safe(row['date'])}</td><td>{safe(row['tech_name'])}</td><td>{int(row['items_count'])}</td><td><span style='background:#27ae60;color:#fff;padding:2px 9px;border-radius:12px;font-size:11px;'>{int(row['ok_count'])}</span></td><td><span style='background:#e67e22;color:#fff;padding:2px 9px;border-radius:12px;font-size:11px;'>{int(row['fix_count'])}</span></td><td style='text-align:right;'>{photos_html}</td></tr>"

        logo_left_html  = render_img(l_logo_b64, "height:70px;object-fit:contain;", '<div style="width:70px;"></div>')
        logo_right_html = render_img(r_logo_b64, "height:70px;object-fit:contain;", '<div style="width:70px;"></div>')

        maint_table = (f"<table><thead><tr><th>الرقم</th><th>التاريخ</th><th>القسم</th><th>الموقع</th><th>الوصف</th><th>الحالة</th><th>الفني</th><th>قبل</th><th>بعد</th></tr></thead><tbody>{maint_rows}</tbody></table>" if total > 0 else '<div class="no-data">لا توجد بلاغات في هذه الفترة</div>')
        clean_table = (f"<table><thead><tr><th>الرقم</th><th>التاريخ</th><th>المنطقة</th><th>نوع التنظيف</th><th>المنفذ</th><th>قبل</th><th>بعد</th></tr></thead><tbody>{clean_rows}</tbody></table>" if clean_count > 0 else '<div class="no-data">لا توجد مهام نظافة في هذه الفترة</div>')
        daily_table = (f"<table><thead><tr><th>الرقم</th><th>التاريخ</th><th>المنفذ</th><th>عدد البنود</th><th>سليم</th><th>يحتاج صيانة</th><th>الصور / Photos</th></tr></thead><tbody>{daily_rows}</tbody></table>" if daily_count > 0 else '<div class="no-data">لا توجد جولات تفقدية في هذه الفترة</div>')

        monthly_html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar"><head><meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Amiri:wght@400;700&family=Cairo:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:'Cairo',sans-serif;background:#ffffff;color:#1a1a2e;}}
.page{{max-width:980px;margin:0 auto;background:#ffffff;border:1px solid #dde3ec;box-shadow:0 4px 24px rgba(0,0,0,0.10);}}
.header{{background:#ffffff;display:flex;flex-direction:column;border-bottom:1px solid #dde3ec;}}
.header-top{{display:flex;align-items:center;justify-content:space-between;padding:20px 28px;gap:16px;}}
.header-center{{text-align:center;flex:1;}}
.header-center .org{{font-family:'Amiri',serif;font-size:17px;color:#c9aa5f;letter-spacing:1px;margin-bottom:6px;}}
.header-center .title{{font-size:22px;font-weight:700;color:#1b3a6b;margin-bottom:6px;}}
.header-center .period{{font-size:12px;color:#5a6a82;background:#ffffff;display:inline-block;padding:3px 16px;border-radius:20px;border:1px solid #dde3ec;}}
.header-stripe{{height:5px;background:linear-gradient(90deg,#1b3a6b,#c9aa5f,#e8c97a,#c9aa5f,#1b3a6b);}}
.meta-bar{{background:#ffffff;border-bottom:1px solid #dde3ec;padding:10px 28px;display:flex;gap:28px;flex-wrap:wrap;}}
.meta-item{{font-size:12px;color:#5a6a82;}}
.meta-item strong{{color:#1b3a6b;font-weight:600;}}
.body{{padding:24px 28px;}}
.stats{{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:26px;}}
.stat-card{{background:#ffffff;border:1px solid #dde3ec;border-top:4px solid #c9aa5f;border-radius:8px;padding:14px 8px;text-align:center;}}
.stat-card .num{{font-size:32px;font-weight:700;color:#1b3a6b;line-height:1;}}
.stat-card .num.green{{color:#27ae60;}}
.stat-card .num.orange{{color:#e67e22;}}
.stat-card .num.blue{{color:#2980b9;}}
.stat-card .num.purple{{color:#8e44ad;}}
.stat-card .lbl{{font-size:10px;color:#8a96aa;margin-top:6px;font-weight:600;}}
.section-title{{font-size:13px;font-weight:700;color:#1b3a6b;background:#ffffff;border-right:4px solid #c9aa5f;padding:8px 14px;border-radius:4px;margin:22px 0 12px;}}
table{{width:100%;border-collapse:collapse;font-size:12.5px;}}
th{{background:#1b3a6b;color:#ffffff;padding:10px 8px;text-align:center;border:1px solid #1b3a6b;}}
td{{padding:9px 8px;text-align:center;border:1px solid #e8ecf2;color:#333d4d;}}
.divider{{border:none;border-top:1px solid #e8ecf2;margin:22px 0;}}
.footer{{background:#ffffff;border-top:2px solid #dde3ec;padding:18px 28px;display:flex;justify-content:space-between;align-items:center;}}
.footer-info{{font-size:12px;color:#5a6a82;line-height:1.9;}}
.footer-info strong{{color:#1b3a6b;}}
.stamp{{border:2.5px solid #c9aa5f;border-radius:50%;width:82px;height:82px;display:flex;flex-direction:column;align-items:center;justify-content:center;color:#c9aa5f;font-size:10px;font-weight:700;text-align:center;transform:rotate(-10deg);background:#fff;}}
.confidential{{text-align:center;font-size:10px;color:#aab;padding:8px;border-top:1px solid #e8ecf2;letter-spacing:2px;background:#ffffff;}}
.no-data{{text-align:center;color:#aab;padding:22px;font-style:italic;border:1px dashed #dde3ec;border-radius:6px;background:#ffffff;}}
</style></head>
<body><div class="page">
<div class="header"><div class="header-top">{logo_right_html}<div class="header-center"><div class="org">{safe(org_name)}</div><div class="title">التقرير الشهري للمرافق</div><div class="period">الفترة: {date_from.strftime('%Y-%m-%d')} — {date_to.strftime('%Y-%m-%d')}</div></div>{logo_left_html}</div><div class="header-stripe"></div></div>
<div class="meta-bar"><div class="meta-item">📅 <strong>تاريخ الإصدار:</strong> {now_local().strftime('%Y-%m-%d %H:%M')}</div><div class="meta-item">👤 <strong>أُعدَّ بواسطة:</strong> {safe(st.session_state.username)}</div><div class="meta-item">📊 <strong>نسبة الإنجاز:</strong> {pct}</div></div>
<div class="body">
<div class="stats">
<div class="stat-card"><div class="num">{total}</div><div class="lbl">إجمالي البلاغات</div></div>
<div class="stat-card"><div class="num green">{done_count}</div><div class="lbl">تم الإصلاح</div></div>
<div class="stat-card"><div class="num orange">{pend_count}</div><div class="lbl">قيد الانتظار</div></div>
<div class="stat-card"><div class="num blue">{clean_count}</div><div class="lbl">مهام النظافة</div></div>
<div class="stat-card"><div class="num purple">{daily_count}</div><div class="lbl">الجولات التفقدية</div></div>
</div>
<hr class="divider"><div class="section-title">🛠️ جدول بلاغات الصيانة</div>{maint_table}
<hr class="divider"><div class="section-title">🧹 جدول مهام النظافة</div>{clean_table}
<hr class="divider"><div class="section-title">✅ جدول الجولات اليومية للتفقّد</div>{daily_table}
</div>
<div class="footer"><div class="footer-info"><strong>تاريخ الإصدار:</strong> {now_local().strftime('%Y-%m-%d %H:%M')}<br><strong>أُعدَّ بواسطة:</strong> {safe(st.session_state.username)}</div><div class="stamp">تم<br>الاعتماد<br>✦</div><div class="footer-info" style="text-align:left;"><strong>نسبة الإنجاز:</strong> {pct}<br><strong>إجمالي البلاغات:</strong> {total}</div></div>
<div class="confidential">{safe(report_footer)}</div>
</div></body></html>"""

        components.html(monthly_html, height=1100, scrolling=True)
        st.write("")

        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                t("download_monthly"),
                monthly_html,
                file_name=f"Monthly_{date_from}_{date_to}.html",
                mime="text/html",
                use_container_width=True
            )
        with col_dl2:
            st.caption(t("excel_zip_info"))
            with st.spinner("جارٍ تحضير ملف ZIP..."):
                _zip_data = create_excel_zip(df_m, df_c, df_d_items, date_from, date_to)
            if _zip_data:
                st.download_button(
                    t("download_excel_zip"),
                    _zip_data,
                    file_name=f"Monthly_Report_{date_from}_to_{date_to}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    type="primary"
                )

# ===================== إدارة المستخدمين (مدير فقط) =====================

elif choice == t("m_users"):
    require_admin()
    st.header(t("m_users"))

    with get_db() as conn:
        users_df = pd.read_sql_query("SELECT username, role FROM users", conn)

    st.subheader(t("current_users"))
    st.dataframe(users_df, use_container_width=True, hide_index=True)

    st.divider()

    st.subheader(t("add_user_h"))
    with st.form("add_user_form"):
        new_user = st.text_input(t("username"))
        new_pass = st.text_input(t("password"), type="password")
        new_role = st.selectbox(t("role"), SELECTABLE_ROLES, format_func=tr_display)
        if st.form_submit_button(t("add_btn"), use_container_width=True):
            if not new_user or not new_pass:
                st.warning(t("fill_all"))
            elif len(new_pass) < MIN_PASSWORD_LEN:
                st.warning(t("pwd_too_short", MIN_PASSWORD_LEN))
            else:
                try:
                    with get_db() as conn:
                        conn.execute(
                            "INSERT INTO users VALUES (?,?,?)",
                            (new_user, hash_password(new_pass), new_role)
                        )
                    st.success(t("user_added", new_user))
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error(t("user_exists"))

    st.divider()

    st.subheader(t("delete_user_h"))
    deletable = [u for u in users_df['username'].tolist() if u != st.session_state.username]
    if deletable:
        with st.form("delete_user_form"):
            del_user = st.selectbox(t("select_to_delete"), deletable)
            if st.form_submit_button(t("delete_btn"), use_container_width=True):
                with get_db() as conn:
                    conn.execute("DELETE FROM users WHERE username=?", (del_user,))
                st.success(t("user_deleted", del_user))
                st.rerun()
    else:
        st.info(t("no_other_users"))

    st.divider()

    st.subheader(t("change_pwd_h"))
    with st.form("change_pass_form"):
        chg_user = st.selectbox(t("select_user"), users_df['username'].tolist(), key="chg_user")
        chg_pass = st.text_input(t("new_pwd"), type="password")
        if st.form_submit_button(t("update_btn"), use_container_width=True):
            if not chg_pass:
                st.warning(t("enter_new_pwd"))
            elif len(chg_pass) < MIN_PASSWORD_LEN:
                st.warning(t("pwd_too_short", MIN_PASSWORD_LEN))
            else:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE users SET password=? WHERE username=?",
                        (hash_password(chg_pass), chg_user)
                    )
                st.success(t("pwd_updated", chg_user))

    st.divider()

    # ============ تعديل دور المستخدم ============
    st.subheader(t("edit_role_h"))
    _editable_users = [u for u in users_df['username'].tolist() if u != st.session_state.username]
    if not _editable_users:
        st.info(t("cant_change_self"))
    else:
        with st.form("edit_role_form"):
            er_user = st.selectbox(t("select_user"), _editable_users, key="er_user")
            # عرض الدور الحالي
            _current = users_df[users_df['username'] == er_user]['role'].iloc[0]
            st.caption(f"{t('current_role')} **{tr_display(_current)}**")
            er_role = st.selectbox(
                t("new_role_label"),
                SELECTABLE_ROLES,
                format_func=tr_display,
                key="er_role"
            )
            if st.form_submit_button(t("update_btn"), use_container_width=True):
                with get_db() as conn:
                    conn.execute(
                        "UPDATE users SET role=? WHERE username=?",
                        (er_role, er_user)
                    )
                st.success(t("role_updated", er_user, tr_display(er_role)))
                st.rerun()

    st.divider()

    # ============ المنطقة الزمنية ============
    st.subheader(t("tz_h"))

    _current_tz = get_setting("timezone", "Asia/Riyadh")
    _tz_keys = list(TIMEZONES.keys())
    try:
        _tz_idx = _tz_keys.index(_current_tz)
    except ValueError:
        _tz_idx = 0

    _selected_tz = st.selectbox(
        t("tz_label"),
        options=_tz_keys,
        format_func=lambda k: TIMEZONES[k],
        index=_tz_idx,
        key="tz_select"
    )
    st.caption(f"{t('tz_current')} **{now_local().strftime('%Y-%m-%d %H:%M:%S')}**")

    if _selected_tz != _current_tz:
        if st.button(t("update_btn"), key="tz_update_btn", use_container_width=True):
            set_setting("timezone", _selected_tz)
            st.success(t("tz_updated", TIMEZONES[_selected_tz]))
            st.rerun()

    st.divider()

    # ============ النسخة الاحتياطية ============
    st.subheader(t("backup_h"))
    st.info(t("backup_info"))

    try:
        import os as _os
        db_size = _os.path.getsize(DB_NAME)
        db_mtime = datetime.fromtimestamp(_os.path.getmtime(DB_NAME))
        with get_db() as _conn:
            total_records = (
                _conn.execute("SELECT COUNT(*) FROM maintenance").fetchone()[0] +
                _conn.execute("SELECT COUNT(*) FROM cleaning").fetchone()[0] +
                _conn.execute("SELECT COUNT(*) FROM daily_checks").fetchone()[0]
            )

        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric(t("backup_size"), f"{db_size/1024:.1f} KB")
        with c2:
            st.metric(t("backup_records"), total_records)
        with c3:
            st.metric(t("backup_last_modified"), db_mtime.strftime("%Y-%m-%d %H:%M"))

        with open(DB_NAME, "rb") as _f:
            db_bytes = _f.read()

        st.download_button(
            label=t("backup_btn"),
            data=db_bytes,
            file_name=f"taqyeem_backup_{now_local().strftime('%Y%m%d_%H%M%S')}.db",
            mime="application/x-sqlite3",
            use_container_width=True,
            type="primary"
        )
    except Exception as _e:
        st.error(f"⚠️ {_e}")

# ===================== زر تسجيل الخروج =====================

if st.sidebar.button(t("logout")):
    st.session_state.clear()
    st.rerun()
